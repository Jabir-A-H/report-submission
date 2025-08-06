from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    jsonify,  # type: ignore
    send_file,  # type: ignore
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager,
    login_user,  # type: ignore
    login_required,  # type: ignore
    logout_user,
    current_user,
    UserMixin,
)
from werkzeug.security import generate_password_hash, check_password_hash
import os
import re
import unicodedata
from pathlib import Path
from dotenv import load_dotenv
from flask_migrate import Migrate
import pandas as pd
import io
from datetime import datetime

load_dotenv()

# --- Initialize Flask App and Configurations ---
BASE_DIR = Path(os.getcwd()).resolve()
INSTANCE_DIR = BASE_DIR / "instance"
INSTANCE_DIR.mkdir(exist_ok=True)
DEFAULT_DB_PATH = INSTANCE_DIR / "reports.db"

db_uri = os.getenv("SQLALCHEMY_DATABASE_URI")
if not db_uri:
    db_uri = f"sqlite:///{DEFAULT_DB_PATH}"

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY")
app.config["SQLALCHEMY_DATABASE_URI"] = db_uri
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"  # type: ignore
migrate = Migrate(app, db)


# Normalize Unicode to NFC and strip whitespace
def normalize_cat(s):
    return unicodedata.normalize("NFC", s).strip()


def slugify(s):
    s = normalize_cat(s)
    s = s.lower()
    s = s.replace(" ", "-")
    for ch in ["(", ")", "্", "়", "’", '"', "'"]:
        s = s.replace(ch, "")
    s = re.sub(r"[^\w\-]+", "", s)
    return s


# DRY helper for slug mappings
def make_slugs(categories):
    norm_categories = [normalize_cat(cat) for cat in categories]
    slugs = [slugify(cat) for cat in norm_categories]
    slug_to_cat = {slug: cat for slug, cat in zip(slugs, norm_categories)}
    cat_to_slug = {cat: slug for cat, slug in zip(norm_categories, slugs)}
    return norm_categories, slug_to_cat, cat_to_slug


# --- Helper: Populate Categories for New Report ---
def populate_categories_for_report(report_id):
    from sqlalchemy.exc import IntegrityError

    # Course Categories
    course_categories = [
        "বিশিষ্টদের",
        "সাধারণদের",
        "কর্মীদের",
        "ইউনিট সভানেত্রী",
        "অগ্রসরদের",
        "শিশু- তা'লিমুল কুরআন",
        "নিরক্ষর- তা'লিমুস সলাত",
    ]
    for cat in course_categories:
        norm_cat = normalize_cat(cat)
        if not ReportCourse.query.filter_by(
            category=norm_cat, report_id=report_id
        ).first():
            try:
                db.session.add(
                    ReportCourse(category=norm_cat, number=0, report_id=report_id)
                )
                db.session.commit()
            except IntegrityError:
                db.session.rollback()

    # Organizational Categories
    org_categories = [
        "দাওয়াত দান",
        "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
        "সহযোগী হয়েছে",
        "সম্মতি দিয়েছেন",
        "সক্রিয় সহযোগী",
        "কর্মী",
        "রুকন",
        "দাওয়াতী ইউনিট",
        "ইউনিট",
        "সূধী",
        "এককালীন",
        "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
        "বই বিলি",
        "বই বিক্রি",
    ]
    for cat in org_categories:
        norm_cat = normalize_cat(cat)
        if not ReportOrganizational.query.filter_by(
            category=norm_cat, report_id=report_id
        ).first():
            try:
                db.session.add(
                    ReportOrganizational(
                        category=norm_cat, number=0, report_id=report_id
                    )
                )
                db.session.commit()
            except IntegrityError:
                db.session.rollback()

    # Personal Activities Categories
    personal_categories = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
    for cat in personal_categories:
        norm_cat = normalize_cat(cat)
        if not ReportPersonal.query.filter_by(
            category=norm_cat, report_id=report_id
        ).first():
            try:
                db.session.add(ReportPersonal(category=norm_cat, report_id=report_id))
                db.session.commit()
            except IntegrityError:
                db.session.rollback()

    # Meeting Categories
    meeting_categories = [
        "কমিটি বৈঠক হয়েছে",
        "মুয়াল্লিমাদের নিয়ে বৈঠক",
        "Committee Orientation",
        "Muallima Orientation",
    ]
    for cat in meeting_categories:
        norm_cat = normalize_cat(cat)
        if not ReportMeeting.query.filter_by(
            category=norm_cat, report_id=report_id
        ).first():
            try:
                db.session.add(ReportMeeting(category=norm_cat, report_id=report_id))
                db.session.commit()
            except IntegrityError:
                db.session.rollback()

    # Extra Activity Categories
    extra_categories = [
        "মক্তব সংখ্যা",
        "মক্তব বৃদ্ধি",
        "মহানগরী পরিচালিত",
        "স্থানীয়ভাবে পরিচালিত",
        "মহানগরীর সফর",
        "থানা কমিটির সফর",
        "থানা প্রতিনিধির সফর",
        "ওয়ার্ড প্রতিনিধির সফর",
    ]
    for cat in extra_categories:
        norm_cat = normalize_cat(cat)
        if not ReportExtra.query.filter_by(
            category=norm_cat, report_id=report_id
        ).first():
            try:
                db.session.add(
                    ReportExtra(category=norm_cat, number=0, report_id=report_id)
                )
                db.session.commit()
            except IntegrityError:
                db.session.rollback()


# --- Models ---


class Zone(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    users = db.relationship("User", backref="zone", lazy=True)
    reports = db.relationship("Report", backref="zone", lazy=True)


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(3), unique=True, nullable=False)  # 3-digit user ID
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(10), nullable=False, default="user")  # 'user' or 'admin'
    active = db.Column(db.Boolean, default=False)
    zone_id = db.Column(db.Integer, db.ForeignKey("zone.id"), nullable=False)

    def get_id(self):
        return str(self.id)


class Report(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    zone_id = db.Column(db.Integer, db.ForeignKey("zone.id"), nullable=False)
    month = db.Column(db.Integer, nullable=True)  # Only for মাসিক
    year = db.Column(db.Integer, nullable=False)
    report_type = db.Column(db.String(20), nullable=False)  # মাসিক, ত্রৈমাসিক, ...
    header = db.relationship("ReportHeader", uselist=False, backref="report")
    courses = db.relationship("ReportCourse", backref="report", lazy=True)
    organizational = db.relationship(
        "ReportOrganizational", backref="report", lazy=True
    )
    personal = db.relationship("ReportPersonal", backref="report", lazy=True)
    meetings = db.relationship("ReportMeeting", backref="report", lazy=True)
    extras = db.relationship("ReportExtra", backref="report", lazy=True)
    comments = db.relationship("ReportComment", uselist=False, backref="report")


class ReportHeader(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    responsible_name = db.Column(db.String(100), nullable=False)
    thana = db.Column(db.String(100), nullable=False)
    ward = db.Column(db.Integer, nullable=False)
    total_muallima = db.Column(db.Integer, default=0, nullable=False)
    muallima_increase = db.Column(db.Integer, default=0, nullable=False)
    muallima_decrease = db.Column(db.Integer, default=0, nullable=False)
    certified_muallima = db.Column(db.Integer, default=0, nullable=False)
    certified_muallima_taking_classes = db.Column(db.Integer, default=0, nullable=False)
    trained_muallima = db.Column(db.Integer, default=0, nullable=False)
    trained_muallima_taking_classes = db.Column(db.Integer, default=0, nullable=False)
    total_unit = db.Column(db.Integer, default=0, nullable=False)
    units_with_muallima = db.Column(db.Integer, default=0, nullable=False)


class ReportCourse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    number = db.Column(db.Integer, default=0, nullable=False)
    increase = db.Column(db.Integer, default=0, nullable=False)
    decrease = db.Column(db.Integer, default=0, nullable=False)
    sessions = db.Column(db.Integer, default=0, nullable=False)
    students = db.Column(db.Integer, default=0, nullable=False)
    attendance = db.Column(db.Integer, default=0, nullable=False)
    status_board = db.Column(db.Integer, default=0, nullable=False)
    status_qayda = db.Column(db.Integer, default=0, nullable=False)
    status_ampara = db.Column(db.Integer, default=0, nullable=False)
    status_quran = db.Column(db.Integer, default=0, nullable=False)
    completed = db.Column(db.Integer, default=0, nullable=False)
    correctly_learned = db.Column(db.Integer, default=0, nullable=False)


class ReportOrganizational(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    number = db.Column(db.Integer, default=0, nullable=False)
    increase = db.Column(db.Integer, default=0, nullable=False)
    amount = db.Column(db.Integer, nullable=True)
    comments = db.Column(db.String(200), nullable=True)


class ReportPersonal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    teaching = db.Column(db.Integer, default=0, nullable=False)
    learning = db.Column(db.Integer, default=0, nullable=False)
    olama_invited = db.Column(db.Integer, default=0, nullable=False)
    became_shohojogi = db.Column(db.Integer, default=0, nullable=False)
    became_sokrio_shohojogi = db.Column(db.Integer, default=0, nullable=False)
    became_kormi = db.Column(db.Integer, default=0, nullable=False)
    became_rukon = db.Column(db.Integer, default=0, nullable=False)


class ReportMeeting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    city_count = db.Column(db.Integer, default=0, nullable=False)
    city_avg_attendance = db.Column(db.Integer, default=0, nullable=False)
    thana_count = db.Column(db.Integer, default=0, nullable=False)
    thana_avg_attendance = db.Column(db.Integer, default=0, nullable=False)
    ward_count = db.Column(db.Integer, default=0, nullable=False)
    ward_avg_attendance = db.Column(db.Integer, default=0, nullable=False)
    comments = db.Column(db.Text, nullable=True)


class ReportExtra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    number = db.Column(db.Integer, default=0, nullable=False)


class ReportComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("report.id"), nullable=False)
    comment = db.Column(db.Text, nullable=True)


class CityReportOverride(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=True)
    report_type = db.Column(db.String(20), nullable=False)
    # Optionally, you can add zone_id if you want per-zone overrides
    # zone_id = db.Column(db.Integer, db.ForeignKey("zone.id"), nullable=True)
    section = db.Column(
        db.String(50), nullable=False
    )  # e.g., 'header', 'organizational', etc.
    field = db.Column(
        db.String(50), nullable=False
    )  # e.g., 'total_muallima', 'comments', etc.
    value = db.Column(db.Text, nullable=True)


# --- Login Manager ---


@login_manager.user_loader  # type: ignore
def load_user(user_id):  # type: ignore
    return db.session.get(User, int(user_id))  # type: ignore


# --- Utility Functions ---


def is_admin():
    return current_user.is_authenticated and current_user.role == "admin"


# --- Utility function to generate next user_id ---
def generate_next_user_id():
    last_user = User.query.order_by(User.user_id.desc()).first()
    if last_user and last_user.user_id.isdigit():
        next_id = int(last_user.user_id) + 1
    else:
        next_id = 100
    return f"{next_id:03d}"


# --- Jinja2 Filters ---


@app.template_filter("month_name")
def month_name(month):
    months = [
        "জানুয়ারি",
        "ফেব্রুয়ারি",
        "মার্চ",
        "এপ্রিল",
        "মে",
        "জুন",
        "জুলাই",
        "আগস্ট",
        "সেপ্টেম্বর",
        "অক্টোবর",
        "নভেম্বর",
        "ডিসেম্বর",
    ]
    try:
        return months[int(month) - 1]
    except (IndexError, ValueError, TypeError):
        return ""


# --- Jinja2 Global: Get Current Report ---
def get_current_report(zone_id, month, year, report_type):
    query = {
        "zone_id": zone_id,
        "year": year,
        "report_type": report_type,
    }
    if report_type == "মাসিক":
        query["month"] = month
    return Report.query.filter_by(**query).first()


app.jinja_env.globals.update(get_current_report=get_current_report)


# --- Routes ---


@app.route("/")
@login_required
def dashboard():
    from datetime import datetime

    month = request.args.get("month")
    year = request.args.get("year")
    report_type = request.args.get("report_type")
    now = datetime.now()
    if not month:
        month = now.month
    if not year:
        year = now.year
    if not report_type:
        report_type = "মাসিক"  # Default report type

    # Define available report types
    report_types = [
        {"value": "মাসিক", "label": "মাসিক"},
        {"value": "ত্রৈমাসিক", "label": "ত্রৈমাসিক"},
        {"value": "ষান্মাসিক", "label": "ষান্মাসিক"},
        {"value": "নয়-মাসিক", "label": "নয়-মাসিক"},
        {"value": "বার্ষিক", "label": "বার্ষিক"},
    ]

    if is_admin():
        reports = Report.query.all()
        city_report_url = url_for("city_report_page")
        return render_template(
            "index_admin.html",
            user=current_user,
            month=month,
            year=year,
            report_type=report_type,
            report_types=report_types,
            reports=reports,
            city_report_url=city_report_url,
        )
    else:
        # Build sections list with completion status and navigation URLs
        section_defs = [
            {
                "name": "মূল তথ্য",
                "icon": "📝",
                "url": url_for("report_header", month=month, year=year),
                "model": ReportHeader,
            },
            {
                "name": "গ্রুপ / কোর্স রিপোর্ট",
                "icon": "📚",
                "url": url_for("report_courses", month=month, year=year),
                "model": ReportCourse,
            },
            {
                "name": "দাওয়াত ও সংগঠন",
                "icon": "🏢",
                "url": url_for("report_organizational", month=month, year=year),
                "model": ReportOrganizational,
            },
            {
                "name": "ব্যক্তিগত উদ্যোগে তা’লীমুল কুরআন",
                "icon": "🗃",
                "url": url_for("report_personal", month=month, year=year),
                "model": ReportPersonal,
            },
            {
                "name": "বৈঠকসমূহ",
                "icon": "🤝",
                "url": url_for("report_meetings", month=month, year=year),
                "model": ReportMeeting,
            },
            {
                "name": "মক্তব ও সফর রিপোর্ট",
                "icon": "✨",
                "url": url_for("report_extras", month=month, year=year),
                "model": ReportExtra,
            },
            {
                "name": "মন্তব্য রিপোর্ট",
                "icon": "💬",
                "url": url_for("report_comments", month=month, year=year),
                "model": ReportComment,
            },
        ]

        # Find the user's zone report for the selected period
        report = Report.query.filter_by(
            zone_id=current_user.zone_id,
            month=month,
            year=year,
        ).first()
        # If a new report was just created, populate its categories
        if report:
            populate_categories_for_report(report.id)

        # Map model class to relationship attribute for uselist=False
        model_to_attr = {
            "ReportHeader": "header",
            "ReportComment": "comments",
        }
        sections = []
        for sdef in section_defs:
            completed = False
            if report:
                model_name = sdef["model"].__name__
                rel_attr = model_to_attr.get(model_name, model_name.lower())
                if hasattr(report, rel_attr):
                    section_obj = getattr(report, rel_attr)
                    # uselist=False: object, uselist=True: list
                    if isinstance(section_obj, list):
                        completed = section_obj and len(section_obj) > 0
                    else:
                        completed = section_obj is not None
            sections.append(
                {
                    "name": sdef["name"],
                    "icon": sdef["icon"],
                    "url": sdef["url"],
                    "completed": completed,
                }
            )

        report_month_year = f"{month}/{year}"

        return render_template(
            "index.html",
            user=current_user,
            month=month,
            year=year,
            sections=sections,
            report_month_year=report_month_year,
        )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identifier = request.form["identifier"]  # Can be email or user_id
        password = request.form["password"]
        user = User.query.filter(
            (User.email == identifier) | (User.user_id == identifier)
        ).first()
        if user and user.active and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid credentials or account not approved.")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]
        zone_id = request.form["zone_id"]
        if User.query.filter_by(email=email).first():
            flash("Email already registered.")
            return redirect(url_for("register"))
        user_id = generate_next_user_id()
        hashed_pw = generate_password_hash(password)
        user = User(
            user_id=user_id,
            name=name,
            email=email,
            password=hashed_pw,
            zone_id=zone_id,
            role="user",
            active=False,
        )
        db.session.add(user)
        db.session.commit()
        flash("Registration successful! Await admin approval.")
        return redirect(url_for("login"))
    zones = Zone.query.all()
    return render_template("register.html", zones=zones)


# --- Admin Management ---


@app.route("/users", methods=["GET", "POST"])
@login_required
def admin_users():
    if not is_admin():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        user_id = request.form["user_id"]
        action = request.form["action"]
        user = db.session.get(User, int(user_id))
        if action == "approve":
            user.active = True
        elif action == "reject":
            db.session.delete(user)
        elif action == "assign_zone":
            zone_id = request.form.get("zone_id")
            if zone_id:
                user.zone_id = int(zone_id)
        db.session.commit()
    users = User.query.all()
    zones = Zone.query.all()
    return render_template("users.html", users=users, zones=zones)


@app.route("/zones", methods=["GET", "POST"])
@login_required
def admin_zones():
    if not is_admin():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        name = request.form.get("name")
        if not name:
            flash("Zone name is required.")
        elif Zone.query.filter_by(name=name).first():
            flash("Zone name already exists.")
        else:
            db.session.add(Zone(name=name))
            db.session.commit()
            flash("Zone added successfully.")
            return redirect(url_for("admin_zones"))
    zones = Zone.query.all()
    return render_template("zones.html", zones=zones)


# --- Delete Zone Route ---
@app.route("/delete_zone/<int:zone_id>", methods=["POST"])
@login_required
def delete_zone(zone_id):
    if not is_admin():
        return redirect(url_for("dashboard"))
    zone = Zone.query.get_or_404(zone_id)
    if zone.users:
        flash("Cannot delete zone: users are assigned to this zone.")
        return redirect(url_for("admin_zones"))
    db.session.delete(zone)
    db.session.commit()
    flash("Zone deleted successfully.")
    return redirect(url_for("admin_zones"))


# --- City Report Page ---
def get_city_report_overrides(year, month, report_type):
    """Fetch all overrides for the given city report period as a dict: {(section, field): value}"""
    query = CityReportOverride.query.filter_by(year=year, report_type=report_type)
    if month:
        query = query.filter_by(month=month)
    else:
        query = query.filter(CityReportOverride.month.is_(None))
    overrides = query.all()
    return {(o.section, o.field): o.value for o in overrides}


def apply_overrides_to_agg(agg_dict, section, overrides):
    """Apply overrides to a dict of aggregated values for a section."""
    for field in agg_dict:
        key = (section, field)
        if key in overrides:
            val = overrides[key]
            # Try to cast to int if original is int, else keep as string
            if isinstance(agg_dict[field], int):
                try:
                    agg_dict[field] = int(val)
                except Exception:
                    agg_dict[field] = val
            else:
                agg_dict[field] = val
    return agg_dict


@app.route("/city_report", methods=["GET", "POST"])
@login_required
def city_report_page():
    if not is_admin():
        return redirect(url_for("dashboard"))
    from datetime import datetime

    month = request.args.get("month")
    year = request.args.get("year")
    report_type = request.args.get("report_type", "মাসিক")
    zone_id = request.args.get("zone_id")
    now = datetime.now()
    if not month and report_type == "মাসিক":
        month = now.month
    if not year:
        year = now.year

    # Category lists and slug mappings (same as report.html)
    course_categories_raw = [
        "বিশিষ্টদের",
        "সাধারণদের",
        "কর্মীদের",
        "ইউনিট সভানেত্রী",
        "অগ্রসরদের",
        "শিশু- তা'লিমুল কুরআন",
        "নিরক্ষর- তা'লিমুস সলাত",
    ]
    org_categories_raw = [
        "দাওয়াত দান",
        "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
        "সহযোগী হয়েছে",
        "সম্মতি দিয়েছেন",
        "সক্রিয় সহযোগী",
        "কর্মী",
        "রুকন",
        "দাওয়াতী ইউনিট",
        "ইউনিট",
        "সূধী",
        "এককালীন",
        "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
        "বই বিলি",
        "বই বিক্রি",
    ]
    personal_categories_raw = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
    meeting_categories_raw = [
        "কমিটি বৈঠক হয়েছে",
        "মুয়াল্লিমাদের নিয়ে বৈঠক",
        "Committee Orientation",
        "Muallima Orientation",
    ]
    extra_categories_raw = [
        "মক্তব সংখ্যা",
        "মক্তব বৃদ্ধি",
        "মহানগরী পরিচালিত",
        "স্থানীয়ভাবে পরিচালিত",
        "মহানগরীর সফর",
        "থানা কমিটির সফর",
        "থানা প্রতিনিধির সফর",
        "ওয়ার্ড প্রতিনিধির সফর",
    ]

    course_categories, cat_to_slug, slug_to_cat = make_slugs(course_categories_raw)
    org_categories, org_cat_to_slug, org_slug_to_cat = make_slugs(org_categories_raw)
    personal_categories, personal_cat_to_slug, personal_slug_to_cat = make_slugs(
        personal_categories_raw
    )
    meeting_categories, meeting_cat_to_slug, meeting_slug_to_cat = make_slugs(
        meeting_categories_raw
    )
    extra_categories, extra_cat_to_slug, extra_slug_to_cat = make_slugs(
        extra_categories_raw
    )

    # --- Aggregation Logic ---
    from sqlalchemy import func

    # Determine which months to include based on report_type
    def get_months(report_type, month):
        if report_type == "মাসিক":
            return [int(month)] if month else []
        elif report_type == "ত্রৈমাসিক":
            return [1, 2, 3]
        elif report_type == "ষান্মাসিক":
            return [1, 2, 3, 4, 5, 6]
        elif report_type == "নয়-মাসিক":
            return [1, 2, 3, 4, 5, 6, 7, 8, 9]
        elif report_type == "বার্ষিক":
            return list(range(1, 13))
        return []

    months = get_months(report_type, month)
    year_int = int(year)
    # If a zone is selected, show only that zone's reports for the period (for any report type)
    if zone_id:
        if report_type == "মাসিক":
            report_query = Report.query.filter_by(
                year=year_int, report_type="মাসিক", zone_id=zone_id
            )
            report_query = report_query.filter(Report.month == int(month))
        else:
            report_query = Report.query.filter_by(
                year=year_int, report_type="মাসিক", zone_id=zone_id
            )
            if months:
                report_query = report_query.filter(Report.month.in_(months))
        reports = report_query.all()
    else:
        # Always aggregate city report from all zones' মাসিক reports for the relevant months
        if report_type == "মাসিক":
            report_query = Report.query.filter_by(year=year_int, report_type="মাসিক")
            report_query = report_query.filter(Report.month == int(month))
        else:
            report_query = Report.query.filter_by(year=year_int, report_type="মাসিক")
            if months:
                report_query = report_query.filter(Report.month.in_(months))
        reports = report_query.all()
    zones = Zone.query.all()

    # --- Fetch overrides for this period ---
    # Only apply overrides for city-wide reports (when zone_id is None)
    if zone_id:
        # For zone-specific reports, don't apply any overrides
        overrides = {}
    else:
        # For city-wide aggregated reports, apply overrides
        overrides = get_city_report_overrides(
            year_int, int(month) if month else None, report_type
        )

    # --- Header Aggregation ---
    header_fields = [
        "ward",
        "total_muallima",
        "muallima_increase",
        "muallima_decrease",
        "certified_muallima",
        "certified_muallima_taking_classes",
        "trained_muallima",
        "trained_muallima_taking_classes",
        "total_unit",
        "units_with_muallima",
    ]
    city_summary = {field: 0 for field in header_fields}
    thana_values = []
    for r in reports:
        if r.header:
            for field in header_fields:
                city_summary[field] += getattr(r.header, field, 0) or 0
            thana_val = getattr(r.header, "thana", None)
            if thana_val is not None:
                try:
                    thana_int = int(thana_val)
                    thana_values.append(thana_int)
                except (ValueError, TypeError):
                    thana_values.append(None)
    city_summary["responsible_name"] = None
    # If all thana_values are integers (not None), sum them; else None
    if thana_values and all(v is not None for v in thana_values):
        city_summary["thana"] = sum(thana_values)
    else:
        city_summary["thana"] = None
    # Apply overrides to header
    city_summary = apply_overrides_to_agg(city_summary, "header", overrides)

    # --- Courses Aggregation ---
    city_courses = []
    for cat in course_categories:
        agg = {"category": cat}
        for field in [
            "number",
            "increase",
            "decrease",
            "sessions",
            "students",
            "attendance",
            "status_board",
            "status_qayda",
            "status_ampara",
            "status_quran",
            "completed",
            "correctly_learned",
        ]:
            agg[field] = 0
        for r in reports:
            for row in r.courses:
                if normalize_cat(row.category) == normalize_cat(cat):
                    for field in agg:
                        if field != "category":
                            agg[field] += getattr(row, field, 0) or 0
        # Apply overrides to this course category
        agg = apply_overrides_to_agg(agg, f"courses:{cat}", overrides)
        city_courses.append(agg)

    # --- Organizational Aggregation ---
    city_organizational = []
    for cat in org_categories:
        agg = {"category": cat, "number": 0, "increase": 0, "amount": 0, "comments": 0}
        found_nonint_comment = False
        for r in reports:
            for row in r.organizational:
                if normalize_cat(row.category) == normalize_cat(cat):
                    agg["number"] += row.number or 0
                    agg["increase"] += row.increase or 0
                    agg["amount"] += row.amount or 0
                    # Comments: sum if integer, else skip
                    if row.comments is not None and not found_nonint_comment:
                        try:
                            agg["comments"] += int(row.comments)
                        except (ValueError, TypeError):
                            found_nonint_comment = True
        if found_nonint_comment:
            agg["comments"] = None
        # Apply overrides to this org category
        agg = apply_overrides_to_agg(agg, f"organizational:{cat}", overrides)
        city_organizational.append(agg)

    # --- Personal Aggregation ---
    city_personal = []
    for cat in personal_categories:
        agg = {
            "category": cat,
            "teaching": 0,
            "learning": 0,
            "olama_invited": 0,
            "became_shohojogi": 0,
            "became_sokrio_shohojogi": 0,
            "became_kormi": 0,
            "became_rukon": 0,
        }
        for r in reports:
            for row in r.personal:
                if normalize_cat(row.category) == normalize_cat(cat):
                    for field in agg:
                        if field != "category":
                            agg[field] += getattr(row, field, 0) or 0
        agg = apply_overrides_to_agg(agg, f"personal:{cat}", overrides)
        city_personal.append(agg)

    # --- Meetings Aggregation ---
    city_meetings = []
    for cat in meeting_categories:
        agg = {
            "category": cat,
            "city_count": 0,
            "city_avg_attendance": 0,
            "thana_count": 0,
            "thana_avg_attendance": 0,
            "ward_count": 0,
            "ward_avg_attendance": 0,
            "comments": 0,
        }
        found_nonint_comment = False
        city_count_sum = 0
        city_att_sum = 0
        thana_count_sum = 0
        thana_att_sum = 0
        ward_count_sum = 0
        ward_att_sum = 0
        n_city = n_thana = n_ward = 0
        for r in reports:
            for row in r.meetings:
                if normalize_cat(row.category) == normalize_cat(cat):
                    city_count_sum += row.city_count or 0
                    if row.city_count:
                        city_att_sum += (row.city_avg_attendance or 0) * row.city_count
                        n_city += row.city_count
                    thana_count_sum += row.thana_count or 0
                    if row.thana_count:
                        thana_att_sum += (
                            row.thana_avg_attendance or 0
                        ) * row.thana_count
                        n_thana += row.thana_count
                    ward_count_sum += row.ward_count or 0
                    if row.ward_count:
                        ward_att_sum += (row.ward_avg_attendance or 0) * row.ward_count
                        n_ward += row.ward_count
                    # Comments: sum if integer, else skip
                    if row.comments is not None and not found_nonint_comment:
                        try:
                            agg["comments"] += int(row.comments)
                        except (ValueError, TypeError):
                            found_nonint_comment = True
        agg["city_count"] = city_count_sum
        agg["city_avg_attendance"] = int(city_att_sum / n_city) if n_city else 0
        agg["thana_count"] = thana_count_sum
        agg["thana_avg_attendance"] = int(thana_att_sum / n_thana) if n_thana else 0
        agg["ward_count"] = ward_count_sum
        agg["ward_avg_attendance"] = int(ward_att_sum / n_ward) if n_ward else 0
        if found_nonint_comment:
            agg["comments"] = None
        agg = apply_overrides_to_agg(agg, f"meetings:{cat}", overrides)
        city_meetings.append(agg)

    # --- Extras Aggregation ---
    city_extras = []
    for cat in extra_categories:
        agg = {"category": cat, "number": 0}
        for r in reports:
            for row in r.extras:
                if normalize_cat(row.category) == normalize_cat(cat):
                    agg["number"] += row.number or 0
        agg = apply_overrides_to_agg(agg, f"extras:{cat}", overrides)
        city_extras.append(agg)

    # --- Comments Aggregation ---
    city_comments = {"comment": ""}
    comment_sum = 0
    found_nonint_comment = False
    comment_strings = []
    for r in reports:
        if r.comments and r.comments.comment is not None:
            try:
                comment_sum += int(r.comments.comment)
            except (ValueError, TypeError):
                found_nonint_comment = True
                if str(r.comments.comment).strip():
                    comment_strings.append(str(r.comments.comment).strip())
    if found_nonint_comment:
        city_comments["comment"] = ", ".join(comment_strings) if comment_strings else ""
    else:
        city_comments["comment"] = comment_sum if comment_sum != 0 else ""
    city_comments = apply_overrides_to_agg(city_comments, "comments", overrides)

    total_zones = len(set(r.zone_id for r in reports))

    # Only show override page URL for city-wide reports (not zone-specific)
    override_page_url = None
    if not zone_id:
        override_page_url = url_for(
            "city_report_override", month=month, year=year, report_type=report_type
        )
    return render_template(
        "city_report.html",
        month=month,
        year=year,
        report_type=report_type,
        course_categories=course_categories,
        org_categories=org_categories,
        personal_categories=personal_categories,
        meeting_categories=meeting_categories,
        extra_categories=extra_categories,
        cat_to_slug=cat_to_slug,
        slug_to_cat=slug_to_cat,
        org_cat_to_slug=org_cat_to_slug,
        org_slug_to_cat=org_slug_to_cat,
        personal_cat_to_slug=personal_cat_to_slug,
        personal_slug_to_cat=personal_slug_to_cat,
        meeting_cat_to_slug=meeting_cat_to_slug,
        meeting_slug_to_cat=meeting_slug_to_cat,
        extra_cat_to_slug=extra_cat_to_slug,
        extra_slug_to_cat=extra_slug_to_cat,
        city_summary=city_summary,
        city_courses=city_courses,
        city_organizational=city_organizational,
        city_personal=city_personal,
        city_meetings=city_meetings,
        city_extras=city_extras,
        city_comments=city_comments,
        reports=reports,
        total_zones=total_zones,
        zones=zones,
        selected_zone_id=zone_id,
        overrides=overrides,
        override_page_url=override_page_url,
    )


# --- City Report Override Admin Page (GET: page, POST: save) ---
@app.route("/city_report/override", methods=["GET", "POST"])
@login_required
def city_report_override():
    if not is_admin():
        return redirect(url_for("dashboard"))
    from datetime import datetime

    if request.method == "POST":
        year = int(request.form.get("year"))
        month = request.form.get("month")
        month = int(month) if month else None
        report_type = request.form.get("report_type")
        section = request.form.get("section")
        field = request.form.get("field")
        value = request.form.get("value")
        # Upsert override
        override = CityReportOverride.query.filter_by(
            year=year,
            month=month,
            report_type=report_type,
            section=section,
            field=field,
        ).first()
        if override:
            override.value = value
        else:
            override = CityReportOverride(
                year=year,
                month=month,
                report_type=report_type,
                section=section,
                field=field,
                value=value,
            )
            db.session.add(override)
        db.session.commit()
        # For AJAX: return OK, else redirect for normal POST
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return ("OK", 200)
        # For normal POST, reload the page with same params
        return redirect(
            url_for(
                "city_report_override", year=year, month=month, report_type=report_type
            )
        )

    # GET: show the override page
    month = request.args.get("month")
    year = request.args.get("year")
    report_type = request.args.get("report_type", "মাসিক")
    now = datetime.now()
    if not month and report_type == "মাসিক":
        month = now.month
    if not year:
        year = now.year
    # Category lists and slug mappings (same as city_report_page)
    course_categories_raw = [
        "বিশিষ্টদের",
        "সাধারণদের",
        "কর্মীদের",
        "ইউনিট সভানেত্রী",
        "অগ্রসরদের",
        "শিশু- তা'লিমুল কুরআন",
        "নিরক্ষর- তা'লিমুস সলাত",
    ]
    org_categories_raw = [
        "দাওয়াত দান",
        "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
        "সহযোগী হয়েছে",
        "সম্মতি দিয়েছেন",
        "সক্রিয় সহযোগী",
        "কর্মী",
        "রুকন",
        "দাওয়াতী ইউনিট",
        "ইউনিট",
        "সূধী",
        "এককালীন",
        "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
        "বই বিলি",
        "বই বিক্রি",
    ]
    personal_categories_raw = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
    meeting_categories_raw = [
        "কমিটি বৈঠক হয়েছে",
        "মুয়াল্লিমাদের নিয়ে বৈঠক",
        "Committee Orientation",
        "Muallima Orientation",
    ]
    extra_categories_raw = [
        "মক্তব সংখ্যা",
        "মক্তব বৃদ্ধি",
        "মহানগরী পরিচালিত",
        "স্থানীয়ভাবে পরিচালিত",
        "মহানগরীর সফর",
        "থানা কমিটির সফর",
        "থানা প্রতিনিধির সফর",
        "ওয়ার্ড প্রতিনিধির সফর",
    ]
    course_categories, cat_to_slug, slug_to_cat = make_slugs(course_categories_raw)
    org_categories, org_cat_to_slug, org_slug_to_cat = make_slugs(org_categories_raw)
    personal_categories, personal_cat_to_slug, personal_slug_to_cat = make_slugs(
        personal_categories_raw
    )
    meeting_categories, meeting_cat_to_slug, meeting_slug_to_cat = make_slugs(
        meeting_categories_raw
    )
    extra_categories, extra_cat_to_slug, extra_slug_to_cat = make_slugs(
        extra_categories_raw
    )

    # Fetch overrides for this period
    query = CityReportOverride.query.filter_by(year=int(year), report_type=report_type)
    if month:
        query = query.filter_by(month=int(month))
    else:
        query = query.filter(CityReportOverride.month.is_(None))
    overrides = query.order_by(
        CityReportOverride.section, CityReportOverride.field
    ).all()

    # --- Aggregation Logic (same as city_report_page, but only for dropdowns/previous values) ---
    # For simplicity, reuse the aggregation logic for the selected period
    # Determine which months to include based on report_type
    def get_months(report_type, month):
        if report_type == "মাসিক":
            return [int(month)] if month else []
        elif report_type == "ত্রৈমাসিক":
            m = int(month) if month else 1
            return [m, m - 1, m - 2]
        elif report_type == "ষান্মাসিক":
            m = int(month) if month else 1
            return [m, m - 1, m - 2, m - 3, m - 4, m - 5]
        elif report_type == "নয়-মাসিক":
            m = int(month) if month else 1
            return [m, m - 1, m - 2, m - 3, m - 4, m - 5, m - 6, m - 7, m - 8]
        elif report_type == "বার্ষিক":
            return list(range(1, 13))
        return []

    months = get_months(report_type, month)
    year_int = int(year)
    from sqlalchemy import func

    # Always aggregate city report from all zones' মাসিক reports for the relevant months
    if report_type == "মাসিক":
        report_query = Report.query.filter_by(
            year=year_int, month=int(month), report_type=report_type
        )
    else:
        report_query = Report.query.filter_by(
            year=year_int, report_type="মাসিক"
        ).filter(Report.month.in_(months))
    reports = report_query.all()

    # --- Header Aggregation ---
    header_fields = [
        "total_unit",
        "total_muallima",
        "muallima_increase",
        "muallima_decrease",
        "certified_muallima",
        "certified_muallima_taking_classes",
        "trained_muallima",
        "trained_muallima_taking_classes",
        "units_with_muallima",
        "ward",
    ]
    city_summary = {field: 0 for field in header_fields}
    thana_values = []
    for r in reports:
        if r.header:
            for field in header_fields:
                city_summary[field] += getattr(r.header, field, 0) or 0
            try:
                thana_val = int(r.header.thana)
            except (ValueError, TypeError, AttributeError):
                thana_val = None
            if thana_val is not None:
                thana_values.append(thana_val)
    city_summary["responsible_name"] = None
    if thana_values and all(v is not None for v in thana_values):
        city_summary["thana"] = sum(thana_values)
    else:
        city_summary["thana"] = None

    # --- Courses Aggregation ---
    city_courses = []
    for cat in course_categories:
        agg = {"category": cat}
        for field in [
            "number",
            "increase",
            "decrease",
            "sessions",
            "students",
            "attendance",
            "status_board",
            "status_qayda",
            "status_ampara",
            "status_quran",
            "completed",
            "correctly_learned",
        ]:
            agg[field] = 0
        for r in reports:
            for row in r.courses:
                if row.category == cat:
                    for field in agg:
                        if field != "category":
                            agg[field] += getattr(row, field, 0) or 0
        city_courses.append(agg)

    # --- Organizational Aggregation ---
    city_organizational = []
    for cat in org_categories:
        agg = {"category": cat, "number": 0, "increase": 0, "amount": 0, "comments": 0}
        found_nonint_comment = False
        for r in reports:
            for row in r.organizational:
                if row.category == cat:
                    agg["number"] += getattr(row, "number", 0) or 0
                    agg["increase"] += getattr(row, "increase", 0) or 0
                    agg["amount"] += getattr(row, "amount", 0) or 0
                    try:
                        agg["comments"] += int(row.comments)
                    except (ValueError, TypeError):
                        found_nonint_comment = True
        if found_nonint_comment:
            agg["comments"] = None
        city_organizational.append(agg)

    # --- Personal Aggregation ---
    city_personal = []
    for cat in personal_categories:
        agg = {
            "category": cat,
            "teaching": 0,
            "learning": 0,
            "olama_invited": 0,
            "became_shohojogi": 0,
            "became_sokrio_shohojogi": 0,
            "became_kormi": 0,
            "became_rukon": 0,
        }
        for r in reports:
            for row in r.personal:
                if row.category == cat:
                    for field in agg:
                        if field != "category":
                            agg[field] += getattr(row, field, 0) or 0
        city_personal.append(agg)

    # --- Meetings Aggregation ---
    city_meetings = []
    for cat in meeting_categories:
        agg = {
            "category": cat,
            "city_count": 0,
            "city_avg_attendance": 0,
            "thana_count": 0,
            "thana_avg_attendance": 0,
            "ward_count": 0,
            "ward_avg_attendance": 0,
            "comments": 0,
        }
        found_nonint_comment = False
        city_count_sum = 0
        city_att_sum = 0
        thana_count_sum = 0
        thana_att_sum = 0
        ward_count_sum = 0
        ward_att_sum = 0
        n_city = n_thana = n_ward = 0
        for r in reports:
            for row in r.meetings:
                if row.category == cat:
                    agg["city_count"] += getattr(row, "city_count", 0) or 0
                    agg["thana_count"] += getattr(row, "thana_count", 0) or 0
                    agg["ward_count"] += getattr(row, "ward_count", 0) or 0
                    try:
                        agg["comments"] += int(row.comments)
                    except (ValueError, TypeError):
                        found_nonint_comment = True
                    if row.city_avg_attendance:
                        city_att_sum += row.city_avg_attendance
                        n_city += 1
                    if row.thana_avg_attendance:
                        thana_att_sum += row.thana_avg_attendance
                        n_thana += 1
                    if row.ward_avg_attendance:
                        ward_att_sum += row.ward_avg_attendance
                        n_ward += 1
        agg["city_avg_attendance"] = int(city_att_sum / n_city) if n_city else 0
        agg["thana_avg_attendance"] = int(thana_att_sum / n_thana) if n_thana else 0
        agg["ward_avg_attendance"] = int(ward_att_sum / n_ward) if n_ward else 0
        if found_nonint_comment:
            agg["comments"] = None
        city_meetings.append(agg)

    # --- Extras Aggregation ---
    city_extras = []
    for cat in extra_categories:
        agg = {"category": cat, "number": 0}
        for r in reports:
            for row in r.extras:
                if row.category == cat:
                    agg["number"] += getattr(row, "number", 0) or 0
        city_extras.append(agg)

    # --- Comments Aggregation ---
    city_comments = {"comment": ""}
    comment_sum = 0
    found_nonint_comment = False
    comment_strings = []
    for r in reports:
        if r.comments and r.comments.comment is not None:
            try:
                comment_sum += int(r.comments.comment)
            except (ValueError, TypeError):
                found_nonint_comment = True
                comment_strings.append(str(r.comments.comment))
    if found_nonint_comment:
        city_comments["comment"] = ", ".join(comment_strings) if comment_strings else ""
    else:
        city_comments["comment"] = comment_sum if comment_sum != 0 else ""

    return render_template(
        "city_report_override.html",
        overrides=overrides,
        year=year,
        month=month,
        report_type=report_type,
        course_categories=course_categories,
        org_categories=org_categories,
        personal_categories=personal_categories,
        meeting_categories=meeting_categories,
        extra_categories=extra_categories,
        cat_to_slug=cat_to_slug,
        slug_to_cat=slug_to_cat,
        org_cat_to_slug=org_cat_to_slug,
        org_slug_to_cat=org_slug_to_cat,
        personal_cat_to_slug=personal_cat_to_slug,
        personal_slug_to_cat=personal_slug_to_cat,
        meeting_cat_to_slug=meeting_cat_to_slug,
        meeting_slug_to_cat=meeting_slug_to_cat,
        extra_cat_to_slug=extra_cat_to_slug,
        extra_slug_to_cat=extra_slug_to_cat,
        city_summary=city_summary,
        city_courses=city_courses,
        city_organizational=city_organizational,
        city_personal=city_personal,
        city_meetings=city_meetings,
        city_extras=city_extras,
        city_comments=city_comments,
    )


# --- Helper: Get Report Period ---
def get_report_period():
    from datetime import datetime

    month = request.args.get("month")
    year = request.args.get("year")
    report_type = request.args.get("report_type", "মাসিক")
    now = datetime.now()
    if not month and report_type == "মাসিক":
        month = now.month
    if not year:
        year = now.year
    return int(month) if month else None, int(year), report_type


# --- DRY Section POST Handler ---
def handle_section_post(report, section_attr, categories, fields):
    section_list = getattr(report, section_attr)
    model = section_list[0].__class__ if section_list else None
    # Normalize all categories for robust matching
    # Always use normalized categories and slugs
    norm_categories = [normalize_cat(cat) for cat in categories]
    slugs = [slugify(cat) for cat in norm_categories]
    slug_to_cat = {slug: cat for slug, cat in zip(slugs, norm_categories)}
    cat_to_slug = {cat: slug for cat, slug in zip(norm_categories, slugs)}
    # Map DB rows by slug for robust matching
    db_rows_by_slug = {slugify(normalize_cat(r.category)): r for r in section_list}
    for slug, cat in slug_to_cat.items():
        row = db_rows_by_slug.get(slug)
        if not row and model:
            print(f"[DEBUG] Creating new row for category: {cat} (slug: {slug})")
            row = model(report_id=report.id, category=cat)
            db.session.add(row)
        if row:
            for field in fields:
                form_key = f"{field}_{slug}"
                value = request.form.get(form_key)
                print(
                    f"[DEBUG] Field: {field}, Slug: {slug}, Form Key: {form_key}, Value: {value}"
                )
                if value is not None:
                    col_type = getattr(row.__class__, field).type
                    if isinstance(col_type, db.Integer):
                        value = int(value) if value else 0
                    elif isinstance(col_type, db.String):
                        value = value.strip() or None
                    print(
                        f"[DEBUG] Setting {field} for category {cat} (slug: {slug}) to {value}"
                    )
                    setattr(row, field, value)
    db.session.commit()
    print("[DEBUG] Database commit complete for section_post.")


@app.route("/report/header", methods=["GET", "POST"])
@login_required
def report_header():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = Report.query.filter_by(
            zone_id=current_user.zone_id,
            month=month,
            year=year,
            report_type=report_type,
        ).first()
        if not report:
            # Create report if not exists
            report = Report(
                zone_id=current_user.zone_id,
                month=month,
                year=year,
                report_type=report_type,
            )
            db.session.add(report)
            db.session.commit()

        if request.method == "POST":
            # Get form data
            responsible_name = request.form.get("responsible_name")
            thana = request.form.get("thana")
            ward = request.form.get("ward")
            total_muallima = request.form.get("total_muallima", type=int)
            muallima_increase = request.form.get("muallima_increase", type=int)
            muallima_decrease = request.form.get("muallima_decrease", type=int)
            certified_muallima = request.form.get("certified_muallima", type=int)
            certified_muallima_taking_classes = request.form.get(
                "certified_muallima_taking_classes", type=int
            )
            trained_muallima = request.form.get("trained_muallima", type=int)
            trained_muallima_taking_classes = request.form.get(
                "trained_muallima_taking_classes", type=int
            )
            total_unit = request.form.get("total_unit", type=int)
            units_with_muallima = request.form.get("units_with_muallima", type=int)

            # Validate required fields
            if not responsible_name or not thana or not ward:
                error = "সব আবশ্যক ঘর পূরণ করুন।"
            else:
                # Create or update ReportHeader
                header = report.header
                if not header:
                    header = ReportHeader(report_id=report.id)
                    db.session.add(header)
                header.responsible_name = responsible_name
                header.thana = thana
                header.ward = ward
                header.total_muallima = total_muallima or 0
                header.muallima_increase = muallima_increase or 0
                header.muallima_decrease = muallima_decrease or 0
                header.certified_muallima = certified_muallima or 0
                header.certified_muallima_taking_classes = (
                    certified_muallima_taking_classes or 0
                )
                header.trained_muallima = trained_muallima or 0
                header.trained_muallima_taking_classes = (
                    trained_muallima_taking_classes or 0
                )
                header.total_unit = total_unit or 0
                header.units_with_muallima = units_with_muallima or 0
                db.session.commit()
                success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
                # Refresh report object
                report = Report.query.filter_by(id=report.id).first()

    return render_template(
        "report/header.html",
        month=month,
        year=year,
        report_type=report_type,
        report=report,
        error=error,
        success=success,
    )


@app.route("/report/courses", methods=["GET", "POST"])
@login_required
def report_courses():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = get_current_report(current_user.zone_id, month, year, report_type)
        course_categories_raw = [
            "বিশিষ্টদের",
            "সাধারণদের",
            "কর্মীদের",
            "ইউনিট সভানেত্রী",
            "অগ্রসরদের",
            "শিশু- তা'লিমুল কুরআন",
            "নিরক্ষর- তা'লিমুস সলাত",
        ]
        course_categories, slug_to_cat, cat_to_slug = make_slugs(course_categories_raw)
        if request.method == "POST" and report:
            fields = [
                "number",
                "increase",
                "decrease",
                "sessions",
                "students",
                "attendance",
                "status_board",
                "status_qayda",
                "status_ampara",
                "status_quran",
                "completed",
                "correctly_learned",
            ]
            handle_section_post(report, "courses", course_categories, fields)
            success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
        return render_template(
            "report/courses.html",
            month=month,
            year=year,
            report_type=report_type,
            report=report,
            error=error,
            success=success,
            course_categories=course_categories,
            slug_to_cat=slug_to_cat,
            cat_to_slug=cat_to_slug,
        )


@app.route("/report/organizational", methods=["GET", "POST"])
@login_required
def report_organizational():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = get_current_report(current_user.zone_id, month, year, report_type)
        org_categories_raw = [
            "দাওয়াত দান",
            "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
            "সহযোগী হয়েছে",
            "সম্মতি দিয়েছেন",
            "সক্রিয় সহযোগী",
            "কর্মী",
            "রুকন",
            "দাওয়াতী ইউনিট",
            "ইউনিট",
            "সূধী",
            "এককালীন",
            "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
            "বই বিলি",
            "বই বিক্রি",
        ]
        org_categories, slug_to_cat, cat_to_slug = make_slugs(org_categories_raw)
        if request.method == "POST" and report:
            fields = ["number", "increase", "amount", "comments"]
            handle_section_post(report, "organizational", org_categories, fields)
            success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
        return render_template(
            "report/organizational.html",
            month=month,
            year=year,
            report_type=report_type,
            report=report,
            error=error,
            success=success,
            org_categories=org_categories,
            slug_to_cat=slug_to_cat,
            cat_to_slug=cat_to_slug,
        )


@app.route("/report/personal", methods=["GET", "POST"])
@login_required
def report_personal():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = get_current_report(current_user.zone_id, month, year, report_type)
        personal_categories_raw = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
        personal_categories, slug_to_cat, cat_to_slug = make_slugs(
            personal_categories_raw
        )
        if request.method == "POST" and report:
            fields = [
                "teaching",
                "learning",
                "olama_invited",
                "became_shohojogi",
                "became_sokrio_shohojogi",
                "became_kormi",
                "became_rukon",
            ]
            handle_section_post(report, "personal", personal_categories, fields)
            success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
        return render_template(
            "report/personal.html",
            month=month,
            year=year,
            report_type=report_type,
            report=report,
            error=error,
            success=success,
            personal_categories=personal_categories,
            slug_to_cat=slug_to_cat,
            cat_to_slug=cat_to_slug,
        )


@app.route("/report/meetings", methods=["GET", "POST"])
@login_required
def report_meetings():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = get_current_report(current_user.zone_id, month, year, report_type)
        meeting_categories_raw = [
            "কমিটি বৈঠক হয়েছে",
            "মুয়াল্লিমাদের নিয়ে বৈঠক",
            "Committee Orientation",
            "Muallima Orientation",
        ]
        meeting_categories, slug_to_cat, cat_to_slug = make_slugs(
            meeting_categories_raw
        )
        if request.method == "POST" and report:
            fields = [
                "city_count",
                "city_avg_attendance",
                "thana_count",
                "thana_avg_attendance",
                "ward_count",
                "ward_avg_attendance",
                "comments",
            ]
            handle_section_post(report, "meetings", meeting_categories, fields)
            success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
        return render_template(
            "report/meetings.html",
            month=month,
            year=year,
            report_type=report_type,
            report=report,
            error=error,
            success=success,
            meeting_categories=meeting_categories,
            slug_to_cat=slug_to_cat,
            cat_to_slug=cat_to_slug,
        )


@app.route("/report/extras", methods=["GET", "POST"])
@login_required
def report_extras():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = get_current_report(current_user.zone_id, month, year, report_type)
        extra_categories_raw = [
            "মক্তব সংখ্যা",
            "মক্তব বৃদ্ধি",
            "মহানগরী পরিচালিত",
            "স্থানীয়ভাবে পরিচালিত",
            "মহানগরীর সফর",
            "থানা কমিটির সফর",
            "থানা প্রতিনিধির সফর",
            "ওয়ার্ড প্রতিনিধির সফর",
        ]
        extra_categories, slug_to_cat, cat_to_slug = make_slugs(extra_categories_raw)
        if request.method == "POST" and report:
            fields = ["number"]
            handle_section_post(report, "extras", extra_categories, fields)
            success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
        return render_template(
            "report/extras.html",
            month=month,
            year=year,
            report_type=report_type,
            report=report,
            error=error,
            success=success,
            extra_categories=extra_categories,
            slug_to_cat=slug_to_cat,
            cat_to_slug=cat_to_slug,
        )


@app.route("/report/comments", methods=["GET", "POST"])
@login_required
def report_comments():
    month, year, report_type = get_report_period()
    report = None
    error = None
    success = None
    if current_user.is_authenticated:
        report = get_current_report(current_user.zone_id, month, year, report_type)
        if request.method == "POST" and report:
            comment = request.form.get("comment", "").strip() or None
            if report.comments:
                report.comments.comment = comment
            else:
                db.session.add(ReportComment(report_id=report.id, comment=comment))
            db.session.commit()
            success = "তথ্য সফলভাবে সংরক্ষণ হয়েছে।"
    return render_template(
        "report/comments.html",
        month=month,
        year=year,
        report_type=report_type,
        report=report,
        error=error,
        success=success,
    )


# --- At a Glance / Summary Report ---


@app.route("/report")
@login_required
def report_summary():
    from datetime import datetime

    report_id = request.args.get("report_id")
    report_type = request.args.get("report_type", "মাসিক")
    month = request.args.get("month")
    year = request.args.get("year")
    now = datetime.now()
    if not month and report_type == "মাসিক":
        month = now.month
    if not year:
        year = now.year

    # Use make_slugs helper for all categories and slug mappings
    course_categories_raw = [
        "বিশিষ্টদের",
        "সাধারণদের",
        "কর্মীদের",
        "ইউনিট সভানেত্রী",
        "অগ্রসরদের",
        "শিশু- তা'লিমুল কুরআন",
        "নিরক্ষর- তা'লিমুস সলাত",
    ]
    org_categories_raw = [
        "দাওয়াত দান",
        "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
        "সহযোগী হয়েছে",
        "সম্মতি দিয়েছেন",
        "সক্রিয় সহযোগী",
        "কর্মী",
        "রুকন",
        "দাওয়াতী ইউনিট",
        "ইউনিট",
        "সূধী",
        "এককালীন",
        "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
        "বই বিলি",
        "বই বিক্রি",
    ]
    personal_categories_raw = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
    meeting_categories_raw = [
        "কমিটি বৈঠক হয়েছে",
        "মুয়াল্লিমাদের নিয়ে বৈঠক",
        "Committee Orientation",
        "Muallima Orientation",
    ]
    extra_categories_raw = [
        "মক্তব সংখ্যা",
        "মক্তব বৃদ্ধি",
        "মহানগরী পরিচালিত",
        "স্থানীয়ভাবে পরিচালিত",
        "মহানগরীর সফর",
        "থানা কমিটির সফর",
        "থানা প্রতিনিধির সফর",
        "ওয়ার্ড প্রতিনিধির সফর",
    ]

    course_categories, cat_to_slug, slug_to_cat = make_slugs(course_categories_raw)
    org_categories, org_cat_to_slug, org_slug_to_cat = make_slugs(org_categories_raw)
    personal_categories, personal_cat_to_slug, personal_slug_to_cat = make_slugs(
        personal_categories_raw
    )
    meeting_categories, meeting_cat_to_slug, meeting_slug_to_cat = make_slugs(
        meeting_categories_raw
    )
    extra_categories, extra_cat_to_slug, extra_slug_to_cat = make_slugs(
        extra_categories_raw
    )

    # --- Aggregation Logic for user/zone reports ---
    def get_months(report_type, month):
        if report_type == "মাসিক":
            return [int(month)] if month else []
        elif report_type == "ত্রৈমাসিক":
            return [1, 2, 3]
        elif report_type == "ষান্মাসিক":
            return [1, 2, 3, 4, 5, 6]
        elif report_type == "নয়-মাসিক":
            return [1, 2, 3, 4, 5, 6, 7, 8, 9]
        elif report_type == "বার্ষিক":
            return list(range(1, 13))
        return []

    months = get_months(report_type, month)
    year_int = int(year)

    # If report_id is provided and user is admin, show that report only (no aggregation)
    if report_id and is_admin():
        report = Report.query.filter_by(id=report_id).first()
        # Optionally, set month/year/report_type from the report for display
        if report:
            month = report.month
            year = report.year
            report_type = report.report_type
        return render_template(
            "report.html",
            report=report,
            report_type=report_type,
            month=month,
            year=year,
            course_categories=course_categories,
            org_categories=org_categories,
            personal_categories=personal_categories,
            meeting_categories=meeting_categories,
            extra_categories=extra_categories,
            cat_to_slug=cat_to_slug,
            slug_to_cat=slug_to_cat,
            org_cat_to_slug=org_cat_to_slug,
            org_slug_to_cat=org_slug_to_cat,
            personal_cat_to_slug=personal_cat_to_slug,
            personal_slug_to_cat=personal_slug_to_cat,
            meeting_cat_to_slug=meeting_cat_to_slug,
            meeting_slug_to_cat=meeting_slug_to_cat,
            extra_cat_to_slug=extra_cat_to_slug,
            extra_slug_to_cat=extra_slug_to_cat,
        )

    # For users: মাসিক (monthly) report is NOT aggregated, just show the single report for that month
    if report_type == "মাসিক":
        report = Report.query.filter_by(
            zone_id=current_user.zone_id,
            year=year_int,
            report_type=report_type,
            month=int(month),
        ).first()
        return render_template(
            "report.html",
            report=report,
            report_type=report_type,
            month=month,
            year=year,
            course_categories=course_categories,
            org_categories=org_categories,
            personal_categories=personal_categories,
            meeting_categories=meeting_categories,
            extra_categories=extra_categories,
            cat_to_slug=cat_to_slug,
            slug_to_cat=slug_to_cat,
            org_cat_to_slug=org_cat_to_slug,
            org_slug_to_cat=org_slug_to_cat,
            personal_cat_to_slug=personal_cat_to_slug,
            personal_slug_to_cat=personal_slug_to_cat,
            meeting_cat_to_slug=meeting_cat_to_slug,
            meeting_slug_to_cat=meeting_slug_to_cat,
            extra_cat_to_slug=extra_cat_to_slug,
            extra_slug_to_cat=extra_slug_to_cat,
        )

    # For other types: aggregate that zone's monthly reports for the required months
    report_query = Report.query.filter_by(
        zone_id=current_user.zone_id, year=year_int, report_type="মাসিক"
    )
    if months:
        report_query = report_query.filter(Report.month.in_(months))
    reports = report_query.all()

    # If no reports, show empty
    if not reports:
        return render_template(
            "report.html",
            report=None,
            report_type=report_type,
            month=month,
            year=year,
            course_categories=course_categories,
            org_categories=org_categories,
            personal_categories=personal_categories,
            meeting_categories=meeting_categories,
            extra_categories=extra_categories,
            cat_to_slug=cat_to_slug,
            slug_to_cat=slug_to_cat,
            org_cat_to_slug=org_cat_to_slug,
            org_slug_to_cat=org_slug_to_cat,
            personal_cat_to_slug=personal_cat_to_slug,
            personal_slug_to_cat=personal_slug_to_cat,
            meeting_cat_to_slug=meeting_cat_to_slug,
            meeting_slug_to_cat=meeting_slug_to_cat,
            extra_cat_to_slug=extra_cat_to_slug,
            extra_slug_to_cat=extra_slug_to_cat,
        )

    # --- Aggregate all sections for the selected period (for non-monthly types) ---
    class AggReport:
        pass

    agg = AggReport()
    # Header
    agg.header = None
    if reports:
        from collections import defaultdict

        header_fields = [
            "total_muallima",
            "muallima_increase",
            "muallima_decrease",
            "certified_muallima",
            "certified_muallima_taking_classes",
            "trained_muallima",
            "trained_muallima_taking_classes",
            "total_unit",
            "units_with_muallima",
            "ward",
        ]
        header_sum = defaultdict(int)
        thana_values = []
        for r in reports:
            if r.header:
                for f in header_fields:
                    header_sum[f] += getattr(r.header, f, 0) or 0
                thana_val = getattr(r.header, "thana", None)
                if thana_val is not None:
                    try:
                        thana_int = int(thana_val)
                        thana_values.append(thana_int)
                    except (ValueError, TypeError):
                        thana_values.append(None)

        class HeaderObj:
            pass

        agg.header = HeaderObj()
        for f in header_fields:
            setattr(agg.header, f, header_sum[f])
        agg.header.responsible_name = None
        if thana_values and all(v is not None for v in thana_values):
            agg.header.thana = sum(thana_values)
        else:
            agg.header.thana = None
    # Courses
    agg.courses = []
    for cat in course_categories:
        agg_row = {"category": cat}
        for field in [
            "number",
            "increase",
            "decrease",
            "sessions",
            "students",
            "attendance",
            "status_board",
            "status_qayda",
            "status_ampara",
            "status_quran",
            "completed",
            "correctly_learned",
        ]:
            agg_row[field] = 0
        for r in reports:
            for row in r.courses:
                if normalize_cat(row.category) == normalize_cat(cat):
                    for field in agg_row:
                        if field != "category":
                            agg_row[field] += getattr(row, field, 0) or 0
        agg.courses.append(agg_row)
    # Organizational
    agg.organizational = []
    for cat in org_categories:
        agg_row = {
            "category": cat,
            "number": 0,
            "increase": 0,
            "amount": 0,
            "comments": 0,
        }
        found_nonint_comment = False
        for r in reports:
            for row in r.organizational:
                if normalize_cat(row.category) == normalize_cat(cat):
                    agg_row["number"] += row.number or 0
                    agg_row["increase"] += row.increase or 0
                    agg_row["amount"] += row.amount or 0
                    # Comments: sum if integer, else skip
                    if row.comments is not None and not found_nonint_comment:
                        try:
                            agg_row["comments"] += int(row.comments)
                        except (ValueError, TypeError):
                            found_nonint_comment = True
        if found_nonint_comment:
            agg_row["comments"] = None
        agg.organizational.append(agg_row)
    # Personal
    agg.personal = []
    for cat in personal_categories:
        agg_row = {
            "category": cat,
            "teaching": 0,
            "learning": 0,
            "olama_invited": 0,
            "became_shohojogi": 0,
            "became_sokrio_shohojogi": 0,
            "became_kormi": 0,
            "became_rukon": 0,
        }
        for r in reports:
            for row in r.personal:
                if normalize_cat(row.category) == normalize_cat(cat):
                    for field in agg_row:
                        if field != "category":
                            agg_row[field] += getattr(row, field, 0) or 0
        agg.personal.append(agg_row)
    # Meetings
    agg.meetings = []
    for cat in meeting_categories:
        agg_row = {
            "category": cat,
            "city_count": 0,
            "city_avg_attendance": 0,
            "thana_count": 0,
            "thana_avg_attendance": 0,
            "ward_count": 0,
            "ward_avg_attendance": 0,
            "comments": 0,
        }
        city_count_sum = 0
        city_att_sum = 0
        thana_count_sum = 0
        thana_att_sum = 0
        ward_count_sum = 0
        ward_att_sum = 0
        n_city = n_thana = n_ward = 0
        found_nonint_comment = False
        for r in reports:
            for row in r.meetings:
                if normalize_cat(row.category) == normalize_cat(cat):
                    city_count_sum += row.city_count or 0
                    if row.city_count:
                        city_att_sum += (row.city_avg_attendance or 0) * row.city_count
                        n_city += row.city_count
                    thana_count_sum += row.thana_count or 0
                    if row.thana_count:
                        thana_att_sum += (
                            row.thana_avg_attendance or 0
                        ) * row.thana_count
                        n_thana += row.thana_count
                    ward_count_sum += row.ward_count or 0
                    if row.ward_count:
                        ward_att_sum += (row.ward_avg_attendance or 0) * row.ward_count
                        n_ward += row.ward_count
                    # Comments: sum if integer, else skip
                    if row.comments is not None and not found_nonint_comment:
                        try:
                            agg_row["comments"] += int(row.comments)
                        except (ValueError, TypeError):
                            found_nonint_comment = True
        agg_row["city_count"] = city_count_sum
        agg_row["city_avg_attendance"] = int(city_att_sum / n_city) if n_city else 0
        agg_row["thana_count"] = thana_count_sum
        agg_row["thana_avg_attendance"] = int(thana_att_sum / n_thana) if n_thana else 0
        agg_row["ward_count"] = ward_count_sum
        agg_row["ward_avg_attendance"] = int(ward_att_sum / n_ward) if n_ward else 0
        if found_nonint_comment:
            agg_row["comments"] = None
        agg.meetings.append(agg_row)
    # Extras
    agg.extras = []
    for cat in extra_categories:
        agg_row = {"category": cat, "number": 0}
        for r in reports:
            for row in r.extras:
                if normalize_cat(row.category) == normalize_cat(cat):
                    agg_row["number"] += row.number or 0
        agg.extras.append(agg_row)

    # Comments: pick first non-empty
    class CommentsObj:
        pass

    agg.comments = CommentsObj()
    comment_sum = 0
    found_nonint_comment = False
    comment_strings = []
    for r in reports:
        if r.comments and r.comments.comment is not None:
            try:
                comment_sum += int(r.comments.comment)
            except (ValueError, TypeError):
                found_nonint_comment = True
                if str(r.comments.comment).strip():
                    comment_strings.append(str(r.comments.comment).strip())
    if found_nonint_comment:
        agg.comments.comment = ", ".join(comment_strings) if comment_strings else ""
    else:
        agg.comments.comment = comment_sum if comment_sum != 0 else ""

    return render_template(
        "report.html",
        report=agg,
        report_type=report_type,
        month=month,
        year=year,
        course_categories=course_categories,
        org_categories=org_categories,
        personal_categories=personal_categories,
        meeting_categories=meeting_categories,
        extra_categories=extra_categories,
        cat_to_slug=cat_to_slug,
        slug_to_cat=slug_to_cat,
        org_cat_to_slug=org_cat_to_slug,
        org_slug_to_cat=org_slug_to_cat,
        personal_cat_to_slug=personal_cat_to_slug,
        personal_slug_to_cat=personal_slug_to_cat,
        meeting_cat_to_slug=meeting_cat_to_slug,
        meeting_slug_to_cat=meeting_slug_to_cat,
        extra_cat_to_slug=extra_cat_to_slug,
        extra_slug_to_cat=extra_slug_to_cat,
    )


# --- Download Features ---


@app.route("/download/excel")
@login_required
def download_excel():
    report_type = request.args.get("report_type", "মাসিক")
    month = request.args.get("month", "জানুয়ারি")
    year = int(request.args.get("year", 2025))
    zone_id = request.args.get("zone_id")  # For admin to download specific zone

    # Get report data based on user role
    if current_user.role == "admin" and zone_id:
        # Admin downloading specific zone report
        zone = Zone.query.get_or_404(zone_id)
        reports = get_reports_for_period(zone_id, report_type, month, year)
        filename = f"Zone_{zone.name}_{report_type}_{year}"
        if report_type == "মাসিক":
            filename += f"_{month}"
    elif current_user.role == "admin":
        # Admin downloading city report (all zones aggregated)
        reports = []
        zones = Zone.query.all()
        for zone in zones:
            zone_reports = get_reports_for_period(zone.id, report_type, month, year)
            reports.extend(zone_reports)
        filename = f"City_Report_{report_type}_{year}"
        if report_type == "মাসিক":
            filename += f"_{month}"
    else:
        # User downloading their zone report
        reports = get_reports_for_period(current_user.zone_id, report_type, month, year)
        filename = f"Zone_{current_user.zone.name}_{report_type}_{year}"
        if report_type == "মাসিক":
            filename += f"_{month}"

    filename += ".xlsx"

    # Create comprehensive Excel data with all sections
    data = []

    # Process report data
    for report_index, report in enumerate(reports):
        # Add zone separator
        if report_index > 0:
            data.append(["", "", "", "", "", "", ""])  # Empty row separator

        if hasattr(report, "zone") and report.zone:
            data.append([f"জোন: {report.zone.name}", "", "", "", "", "", ""])
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Header section
        if report.header:
            data.append(["বিভাগ", "ক্ষেত্র", "মান", "", "", "", ""])
            data.append(
                [
                    "হেডার তথ্য",
                    "দায়িত্বশীলের নাম",
                    report.header.responsible_name or "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                ["হেডার তথ্য", "থানা", str(report.header.thana or ""), "", "", "", ""]
            )
            data.append(
                ["হেডার তথ্য", "ওয়ার্ড", str(report.header.ward or ""), "", "", "", ""]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "মোট মুয়াল্লিমা",
                    str(report.header.total_muallima or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "মুয়াল্লিমা বৃদ্ধি",
                    str(report.header.muallima_increase or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "মুয়াল্লিমা হ্রাস",
                    str(report.header.muallima_decrease or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "সার্টিফিকেটধারী মুয়াল্লিমা",
                    str(report.header.certified_muallima or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "সার্টিফিকেটধারী মুয়াল্লিমা ক্লাস নিচ্ছেন",
                    str(report.header.certified_muallima_taking_classes or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "প্রশিক্ষিত মুয়াল্লিমা",
                    str(report.header.trained_muallima or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "প্রশিক্ষিত মুয়াল্লিমা ক্লাস নিচ্ছেন",
                    str(report.header.trained_muallima_taking_classes or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "মোট ইউনিট",
                    str(report.header.total_unit or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(
                [
                    "হেডার তথ্য",
                    "মুয়াল্লিমা সহ ইউনিট",
                    str(report.header.units_with_muallima or 0),
                    "",
                    "",
                    "",
                    "",
                ]
            )
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Courses section
        if report.courses:
            data.append(
                ["বিভাগ", "গ্রুপ/কোর্স", "সংখ্যা", "বৃদ্ধি", "ঘাটতি", "অধিবেশন", "শিক্ষার্থী"]
            )
            course_categories = [
                "বিশিষ্টদের",
                "সাধারণদের",
                "কর্মীদের",
                "ইউনিট সভানেত্রী",
                "অগ্রসরদের",
                "শিশু- তা'লিমুল কুরআন",
                "নিরক্ষর- তা'লিমুস সলাত",
            ]
            for category in course_categories:
                course_row = next(
                    (c for c in report.courses if c.category == category), None
                )
                data.append(
                    [
                        "গ্রুপ/কোর্স",
                        category,
                        (
                            course_row.number
                            if course_row and course_row.number is not None
                            else 0
                        ),
                        (
                            course_row.increase
                            if course_row and course_row.increase is not None
                            else 0
                        ),
                        (
                            course_row.decrease
                            if course_row and course_row.decrease is not None
                            else 0
                        ),
                        (
                            course_row.sessions
                            if course_row and course_row.sessions is not None
                            else 0
                        ),
                        (
                            course_row.students
                            if course_row and course_row.students is not None
                            else 0
                        ),
                    ]
                )
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Organizational section
        if report.organizational:
            data.append(
                ["বিভাগ", "দাওয়াত ও সংগঠন", "সংখ্যা", "বৃদ্ধি", "পরিমাণ", "মন্তব্য", ""]
            )
            org_categories = [
                "দাওয়াত দান",
                "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
                "সহযোগী হয়েছে",
                "সম্মতি দিয়েছেন",
                "সক্রিয় সহযোগী",
                "কর্মী",
                "রুকন",
                "দাওয়াতী ইউনিট",
                "ইউনিট",
                "সূধী",
                "এককালীন",
                "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
                "বই বিলি",
                "বই বিক্রি",
            ]
            for category in org_categories:
                org_row = next(
                    (o for o in report.organizational if o.category == category), None
                )
                data.append(
                    [
                        "দাওয়াত ও সংগঠন",
                        category,
                        org_row.number if org_row and org_row.number is not None else 0,
                        (
                            org_row.increase
                            if org_row and org_row.increase is not None
                            else 0
                        ),
                        org_row.amount if org_row and org_row.amount is not None else 0,
                        (
                            org_row.comments
                            if org_row and org_row.comments is not None
                            else ""
                        ),
                        "",
                    ]
                )
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Personal section
        if report.personal:
            data.append(
                [
                    "বিভাগ",
                    "ব্যক্তিগত উদ্যোগে তা'লীমুল কুরআন",
                    "কতজন শিখাচ্ছেন",
                    "কতজনকে শিখাচ্ছেন",
                    "ওয়ালামাকে দাওয়াত",
                    "সহযোগী হয়েছেন",
                    "কর্মী হয়েছেন",
                ]
            )
            personal_categories = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
            for category in personal_categories:
                personal_row = next(
                    (p for p in report.personal if p.category == category), None
                )
                data.append(
                    [
                        "ব্যক্তিগত",
                        category,
                        (
                            personal_row.teaching
                            if personal_row and personal_row.teaching is not None
                            else 0
                        ),
                        (
                            personal_row.learning
                            if personal_row and personal_row.learning is not None
                            else 0
                        ),
                        (
                            personal_row.olama_invited
                            if personal_row and personal_row.olama_invited is not None
                            else 0
                        ),
                        (
                            personal_row.became_shohojogi
                            if personal_row
                            and personal_row.became_shohojogi is not None
                            else 0
                        ),
                        (
                            personal_row.became_kormi
                            if personal_row and personal_row.became_kormi is not None
                            else 0
                        ),
                    ]
                )
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Meetings section
        if report.meetings:
            data.append(
                [
                    "বিভাগ",
                    "বৈঠকসমূহ",
                    "মহানগরী কতটি",
                    "মহানগরী গড় উপস্থিতি",
                    "থানা কতটি",
                    "থানা গড় উপস্থিতি",
                    "মন্তব্য",
                ]
            )
            meeting_categories = [
                "কমিটি বৈঠক হয়েছে",
                "মুয়াল্লিমাদের নিয়ে বৈঠক",
                "Committee Orientation",
                "Muallima Orientation",
            ]
            for category in meeting_categories:
                meeting_row = next(
                    (m for m in report.meetings if m.category == category), None
                )
                data.append(
                    [
                        "বৈঠকসমূহ",
                        category,
                        (
                            meeting_row.city_count
                            if meeting_row and meeting_row.city_count is not None
                            else 0
                        ),
                        (
                            meeting_row.city_avg_attendance
                            if meeting_row
                            and meeting_row.city_avg_attendance is not None
                            else 0
                        ),
                        (
                            meeting_row.thana_count
                            if meeting_row and meeting_row.thana_count is not None
                            else 0
                        ),
                        (
                            meeting_row.thana_avg_attendance
                            if meeting_row
                            and meeting_row.thana_avg_attendance is not None
                            else 0
                        ),
                        (
                            meeting_row.comments
                            if meeting_row and meeting_row.comments is not None
                            else ""
                        ),
                    ]
                )
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Extras section
        if report.extras:
            data.append(["বিভাগ", "মক্তব ও সফর রিপোর্ট", "সংখ্যা", "", "", "", ""])
            extra_categories = [
                "মক্তব সংখ্যা",
                "মক্তব বৃদ্ধি",
                "মহানগরী পরিচালিত",
                "স্থানীয়ভাবে পরিচালিত",
                "মহানগরীর সফর",
                "থানা কমিটির সফর",
                "থানা প্রতিনিধির সফর",
                "ওয়ার্ড প্রতিনিধির সফর",
            ]
            for category in extra_categories:
                extra_row = next(
                    (e for e in report.extras if e.category == category), None
                )
                data.append(
                    [
                        "মক্তব ও সফর",
                        category,
                        (
                            extra_row.number
                            if extra_row and extra_row.number is not None
                            else 0
                        ),
                        "",
                        "",
                        "",
                        "",
                    ]
                )
            data.append(["", "", "", "", "", "", ""])  # Empty row

        # Comments section
        if report.comments and report.comments.comment:
            data.append(["বিভাগ", "মন্তব্য", "", "", "", "", ""])
            data.append(["মন্তব্য", report.comments.comment, "", "", "", "", ""])
            data.append(["", "", "", "", "", "", ""])  # Empty row

    # Create DataFrame with flexible column structure
    headers = ["বিভাগ", "বিষয়", "কলাম ১", "কলাম ২", "কলাম ৩", "কলাম ৪", "কলাম ৫"]
    df = pd.DataFrame(data, columns=headers)

    # Create Excel file in memory
    output = io.BytesIO()

    # Download Tiro Bangla font data for Excel (same as PDF)
    bengali_font_name = "Times New Roman"  # Default fallback
    try:
        import urllib.request

        # Try to download Tiro Bangla font info
        font_url = (
            "https://fonts.gstatic.com/s/tirobangla/v6/IuaHaJaHVsk2rGGByzUFTJrmwLs.ttf"
        )
        with urllib.request.urlopen(font_url) as response:
            # If download succeeds, we know the font is available
            bengali_font_name = "Tiro Bangla"
    except:
        # If download fails, use Times New Roman
        bengali_font_name = "Times New Roman"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Report", index=False)

        # Get workbook and worksheet for formatting
        from openpyxl.styles import Font, PatternFill, Border, Side

        workbook = writer.book
        worksheet = writer.sheets["Report"]

        # Create Bengali font style
        bengali_font = Font(name=bengali_font_name, size=11)
        header_font = Font(name=bengali_font_name, size=12, bold=True, color="FFFFFF")

        # Create header fill
        header_fill = PatternFill(
            start_color="4CAF50", end_color="4CAF50", fill_type="solid"
        )

        # Create border
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Format data rows with Bengali font
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = bengali_font
                cell.border = thin_border

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/download/pdf")
@login_required
def download_pdf():
    report_type = request.args.get("report_type", "মাসিক")
    month = request.args.get("month", "জানুয়ারি")
    year = int(request.args.get("year", 2025))
    zone_id = request.args.get("zone_id")

    # Get report data based on user role
    if current_user.role == "admin" and zone_id:
        zone = Zone.query.get_or_404(zone_id)
        reports = get_reports_for_period(zone_id, report_type, month, year)
        title = f"Zone: {zone.name}"
        filename = f"Zone_{zone.name}_{report_type}_{year}"
    elif current_user.role == "admin":
        reports = []
        zones = Zone.query.all()
        for zone in zones:
            zone_reports = get_reports_for_period(zone.id, report_type, month, year)
            reports.extend(zone_reports)
        title = "City Report"
        filename = f"City_Report_{report_type}_{year}"
    else:
        reports = get_reports_for_period(current_user.zone_id, report_type, month, year)
        title = f"Zone: {current_user.zone.name}"
        filename = f"Zone_{current_user.zone.name}_{report_type}_{year}"

    if report_type == "মাসিক":
        title += f" - {month}"
        filename += f"_{month}"

    title += f" {year} - {report_type}"
    filename += ".pdf"

    # Use Playwright for PDF generation
    return generate_pdf_with_playwright(reports, title, filename)


@app.route("/download/city_report_pdf")
@login_required
def download_city_report_pdf():
    """Download aggregated city report as PDF with overrides applied"""
    print(
        f"[DEBUG] City report PDF download route called with params: month={request.args.get('month')}, year={request.args.get('year')}, report_type={request.args.get('report_type')}"
    )

    if not is_admin():
        print("[DEBUG] User is not admin, redirecting to dashboard")
        return redirect(url_for("dashboard"))

    from datetime import datetime

    month = request.args.get("month")
    year = request.args.get("year")
    report_type = request.args.get("report_type", "মাসিক")
    now = datetime.now()

    if not month and report_type == "মাসিক":
        month = now.month
    if not year:
        year = now.year

    year_int = int(year)
    print(
        f"[DEBUG] Final parameters: month={month}, year={year_int}, report_type={report_type}"
    )

    # Generate the aggregated city report data (same logic as city_report_page)
    try:
        print("[DEBUG] Generating city report data...")
        city_data = generate_city_report_data(year_int, month, report_type)
        print(
            f"[DEBUG] City data generated successfully: {len(city_data.get('city_courses', []))} course categories"
        )
    except Exception as e:
        print(f"[DEBUG] Error generating city data: {e}")
        flash(f"Error generating report data: {str(e)}", "error")
        return redirect(url_for("city_report_page"))

    # Create filename
    filename = f"City_Report_{report_type}_{year}"
    if report_type == "মাসিক":
        filename += f"_{month}"
    filename += ".pdf"

    title = f"City Report - {report_type}"
    if report_type == "মাসিক":
        title += f" - {month}"
    title += f" {year}"

    print(f"[DEBUG] About to generate PDF with title: {title}, filename: {filename}")

    # Generate PDF of the aggregated city report
    try:
        return generate_city_report_pdf(city_data, title, filename)
    except Exception as e:
        print(f"[DEBUG] Error in PDF generation: {e}")
        flash(f"PDF generation failed: {str(e)}", "error")
        return redirect(url_for("city_report_page"))


def generate_city_report_data(year, month, report_type):
    """Generate aggregated city report data with overrides applied"""
    # Category lists and slug mappings (same as city_report_page)
    course_categories_raw = [
        "বিশিষ্টদের",
        "সাধারণদের",
        "কর্মীদের",
        "ইউনিট সভানেত্রী",
        "অগ্রসরদের",
        "শিশু- তা'লিমুল কুরআন",
        "নিরক্ষর- তা'লিমুস সলাত",
    ]
    org_categories_raw = [
        "দাওয়াত দান",
        "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
        "সহযোগী হয়েছে",
        "সম্মতি দিয়েছেন",
        "সক্রিয় সহযোগী",
        "কর্মী",
        "রুকন",
        "দাওয়াতী ইউনিট",
        "ইউনিট",
        "সূধী",
        "এককালীন",
        "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
        "বই বিলি",
        "বই বিক্রি",
    ]
    personal_categories_raw = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]
    meeting_categories_raw = [
        "কমিটি বৈঠক হয়েছে",
        "মুয়াল্লিমাদের নিয়ে বৈঠক",
        "Committee Orientation",
        "Muallima Orientation",
    ]
    extra_categories_raw = [
        "মক্তব সংখ্যা",
        "মক্তব বৃদ্ধি",
        "মহানগরী পরিচালিত",
        "স্থানীয়ভাবে পরিচালিত",
        "মহানগরীর সফর",
        "থানা কমিটির সফর",
        "থানা প্রতিনিধির সফর",
        "ওয়ার্ড প্রতিনিধির সফর",
    ]

    course_categories, cat_to_slug, slug_to_cat = make_slugs(course_categories_raw)
    org_categories, org_cat_to_slug, org_slug_to_cat = make_slugs(org_categories_raw)
    personal_categories, personal_cat_to_slug, personal_slug_to_cat = make_slugs(
        personal_categories_raw
    )
    meeting_categories, meeting_cat_to_slug, meeting_slug_to_cat = make_slugs(
        meeting_categories_raw
    )
    extra_categories, extra_cat_to_slug, extra_slug_to_cat = make_slugs(
        extra_categories_raw
    )

    # Determine which months to include based on report_type
    def get_months(report_type, month):
        if report_type == "মাসিক":
            return [int(month)] if month else []
        elif report_type == "ত্রৈমাসিক":
            return [1, 2, 3]
        elif report_type == "ষান্মাসিক":
            return [1, 2, 3, 4, 5, 6]
        elif report_type == "নয়-মাসিক":
            return [1, 2, 3, 4, 5, 6, 7, 8, 9]
        elif report_type == "বার্ষিক":
            return list(range(1, 13))
        return []

    months = get_months(report_type, month)

    # Always aggregate city report from all zones' মাসিক reports for the relevant months
    if report_type == "মাসিক":
        report_query = Report.query.filter_by(year=year, report_type="মাসিক")
        report_query = report_query.filter(Report.month == int(month))
    else:
        report_query = Report.query.filter_by(year=year, report_type="মাসিক")
        if months:
            report_query = report_query.filter(Report.month.in_(months))
    reports = report_query.all()

    # Fetch overrides for this period
    overrides = get_city_report_overrides(
        year, int(month) if month else None, report_type
    )

    # Header Aggregation
    header_fields = [
        "ward",
        "total_muallima",
        "muallima_increase",
        "muallima_decrease",
        "certified_muallima",
        "certified_muallima_taking_classes",
        "trained_muallima",
        "trained_muallima_taking_classes",
        "total_unit",
        "units_with_muallima",
    ]
    city_summary = {field: 0 for field in header_fields}
    thana_values = []
    for r in reports:
        if r.header:
            for field in header_fields:
                city_summary[field] += getattr(r.header, field, 0) or 0
            thana_val = getattr(r.header, "thana", None)
            if thana_val is not None:
                try:
                    thana_int = int(thana_val)
                    thana_values.append(thana_int)
                except (ValueError, TypeError):
                    thana_values.append(None)
    city_summary["responsible_name"] = None
    # If all thana_values are integers (not None), sum them; else None
    if thana_values and all(v is not None for v in thana_values):
        city_summary["thana"] = sum(thana_values)
    else:
        city_summary["thana"] = None
    # Apply overrides to header
    city_summary = apply_overrides_to_agg(city_summary, "header", overrides)

    # Courses Aggregation
    city_courses = []
    for cat in course_categories:
        agg = {"category": cat}
        for field in [
            "number",
            "increase",
            "decrease",
            "sessions",
            "students",
            "attendance",
            "status_board",
            "status_qayda",
            "status_ampara",
            "status_quran",
            "completed",
            "correctly_learned",
        ]:
            agg[field] = 0
        for r in reports:
            for row in r.courses:
                if normalize_cat(row.category) == normalize_cat(cat):
                    for field in agg:
                        if field != "category":
                            agg[field] += getattr(row, field, 0) or 0
        # Apply overrides to this course category
        agg = apply_overrides_to_agg(agg, f"courses:{cat}", overrides)
        city_courses.append(agg)

    # Organizational Aggregation
    city_organizational = []
    for cat in org_categories:
        agg = {"category": cat, "number": 0, "increase": 0, "amount": 0, "comments": 0}
        found_nonint_comment = False
        for r in reports:
            for row in r.organizational:
                if normalize_cat(row.category) == normalize_cat(cat):
                    agg["number"] += row.number or 0
                    agg["increase"] += row.increase or 0
                    agg["amount"] += row.amount or 0
                    # Comments: sum if integer, else skip
                    if row.comments is not None and not found_nonint_comment:
                        try:
                            agg["comments"] += int(row.comments)
                        except (ValueError, TypeError):
                            found_nonint_comment = True
        if found_nonint_comment:
            agg["comments"] = None
        # Apply overrides to this org category
        agg = apply_overrides_to_agg(agg, f"organizational:{cat}", overrides)
        city_organizational.append(agg)

    # Personal Aggregation
    city_personal = []
    for cat in personal_categories:
        agg = {
            "category": cat,
            "teaching": 0,
            "learning": 0,
            "olama_invited": 0,
            "became_shohojogi": 0,
            "became_sokrio_shohojogi": 0,
            "became_kormi": 0,
            "became_rukon": 0,
        }
        for r in reports:
            for row in r.personal:
                if normalize_cat(row.category) == normalize_cat(cat):
                    for field in agg:
                        if field != "category":
                            agg[field] += getattr(row, field, 0) or 0
        agg = apply_overrides_to_agg(agg, f"personal:{cat}", overrides)
        city_personal.append(agg)

    # Meetings Aggregation
    city_meetings = []
    for cat in meeting_categories:
        agg = {
            "category": cat,
            "city_count": 0,
            "city_avg_attendance": 0,
            "thana_count": 0,
            "thana_avg_attendance": 0,
            "ward_count": 0,
            "ward_avg_attendance": 0,
            "comments": 0,
        }
        found_nonint_comment = False
        city_count_sum = 0
        city_att_sum = 0
        thana_count_sum = 0
        thana_att_sum = 0
        ward_count_sum = 0
        ward_att_sum = 0
        n_city = n_thana = n_ward = 0
        for r in reports:
            for row in r.meetings:
                if normalize_cat(row.category) == normalize_cat(cat):
                    city_count_sum += row.city_count or 0
                    if row.city_count:
                        city_att_sum += (row.city_avg_attendance or 0) * row.city_count
                        n_city += row.city_count
                    thana_count_sum += row.thana_count or 0
                    if row.thana_count:
                        thana_att_sum += (
                            row.thana_avg_attendance or 0
                        ) * row.thana_count
                        n_thana += row.thana_count
                    ward_count_sum += row.ward_count or 0
                    if row.ward_count:
                        ward_att_sum += (row.ward_avg_attendance or 0) * row.ward_count
                        n_ward += row.ward_count
                    # Comments: sum if integer, else skip
                    if row.comments is not None and not found_nonint_comment:
                        try:
                            agg["comments"] += int(row.comments)
                        except (ValueError, TypeError):
                            found_nonint_comment = True
        agg["city_count"] = city_count_sum
        agg["city_avg_attendance"] = int(city_att_sum / n_city) if n_city else 0
        agg["thana_count"] = thana_count_sum
        agg["thana_avg_attendance"] = int(thana_att_sum / n_thana) if n_thana else 0
        agg["ward_count"] = ward_count_sum
        agg["ward_avg_attendance"] = int(ward_att_sum / n_ward) if n_ward else 0
        if found_nonint_comment:
            agg["comments"] = None
        agg = apply_overrides_to_agg(agg, f"meetings:{cat}", overrides)
        city_meetings.append(agg)

    # Extras Aggregation
    city_extras = []
    for cat in extra_categories:
        agg = {"category": cat, "number": 0}
        for r in reports:
            for row in r.extras:
                if normalize_cat(row.category) == normalize_cat(cat):
                    agg["number"] += row.number or 0
        agg = apply_overrides_to_agg(agg, f"extras:{cat}", overrides)
        city_extras.append(agg)

    # Comments Aggregation
    city_comments = {"comment": ""}
    comment_sum = 0
    found_nonint_comment = False
    comment_strings = []
    for r in reports:
        if r.comments and r.comments.comment is not None:
            try:
                comment_sum += int(r.comments.comment)
            except (ValueError, TypeError):
                found_nonint_comment = True
                if str(r.comments.comment).strip():
                    comment_strings.append(str(r.comments.comment).strip())
    if found_nonint_comment:
        city_comments["comment"] = ", ".join(comment_strings) if comment_strings else ""
    else:
        city_comments["comment"] = comment_sum if comment_sum != 0 else ""
    city_comments = apply_overrides_to_agg(city_comments, "comments", overrides)

    total_zones = len(set(r.zone_id for r in reports))

    return {
        "month": month,
        "year": year,
        "report_type": report_type,
        "course_categories": course_categories,
        "org_categories": org_categories,
        "personal_categories": personal_categories,
        "meeting_categories": meeting_categories,
        "extra_categories": extra_categories,
        "city_summary": city_summary,
        "city_courses": city_courses,
        "city_organizational": city_organizational,
        "city_personal": city_personal,
        "city_meetings": city_meetings,
        "city_extras": city_extras,
        "city_comments": city_comments,
        "total_zones": total_zones,
        "overrides": overrides,
    }


def generate_city_report_pdf(city_data, title, filename):
    """Generate PDF from aggregated city report data using Playwright"""
    try:
        from playwright.sync_api import sync_playwright
        from datetime import datetime
        import io

        print("[DEBUG] Starting PDF generation with Playwright...")

        # Create HTML content for the city report
        html_content = f"""
        <!DOCTYPE html>
        <html lang="bn">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{title}</title>
            <link href="https://fonts.googleapis.com/css2?family=Tiro+Bangla:ital@0;1&family=Ubuntu:ital,wght@0,300;0,400;0,500;0,700;1,300;1,400;1,500;1,700&display=swap" rel="stylesheet">
            <style>
                body {{
                    font-family: 'Tiro Bangla', sans-serif;
                    margin: 20px;
                    font-size: 11px;
                    line-height: 1.3;
                    color: #333;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 30px;
                    border-bottom: 3px solid #3498db;
                    padding-bottom: 20px;
                }}
                .title {{
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 10px;
                    color: #2c3e50;
                }}
                .subtitle {{
                    font-size: 16px;
                    color: #7f8c8d;
                    margin-bottom: 5px;
                }}
                .report-info {{
                    font-size: 12px;
                    color: #95a5a6;
                }}
                .section {{
                    margin-bottom: 25px;
                    page-break-inside: avoid;
                }}
                .section-title {{
                    font-size: 14px;
                    font-weight: bold;
                    background: linear-gradient(135deg, #3498db, #2980b9);
                    color: white;
                    padding: 10px 15px;
                    border-radius: 6px;
                    margin-bottom: 15px;
                    text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .header-grid {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 15px;
                    margin-bottom: 20px;
                }}
                .header-item {{
                    padding: 8px 12px;
                    background: #f8f9fa;
                    border-left: 4px solid #3498db;
                    border-radius: 4px;
                }}
                .header-label {{
                    font-weight: bold;
                    color: #2c3e50;
                }}
                .header-value {{
                    color: #27ae60;
                    font-weight: 600;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 20px;
                    font-size: 10px;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                    border-radius: 6px;
                    overflow: hidden;
                }}
                th {{
                    background: linear-gradient(135deg, #34495e, #2c3e50);
                    color: white;
                    padding: 10px 6px;
                    text-align: center;
                    font-weight: bold;
                    border: 1px solid #2c3e50;
                    text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
                    font-size: 9px;
                    line-height: 1.2;
                }}
                th.sub-header {{
                    background: linear-gradient(135deg, #5dade2, #3498db);
                    font-size: 8px;
                }}
                td {{
                    padding: 6px;
                    border: 1px solid #bdc3c7;
                    text-align: center;
                    background-color: #ffffff;
                    font-size: 9px;
                }}
                td.category {{
                    text-align: left;
                    font-weight: bold;
                    background-color: #ecf0f1;
                    color: #2c3e50;
                    padding-left: 10px;
                }}
                tr:nth-child(even) td {{
                    background-color: #f8f9fa;
                }}
                tr.totals {{
                    background: linear-gradient(135deg, #85c1e9, #5dade2) !important;
                    font-weight: bold;
                }}
                tr.totals td {{
                    background: linear-gradient(135deg, #85c1e9, #5dade2) !important;
                    color: #1a5490;
                    font-weight: bold;
                }}
                .number {{
                    font-family: 'Ubuntu', sans-serif;
                    text-align: center;
                    font-weight: 600;
                    color: #27ae60;
                }}
                .comments-box {{
                    background: #f8f9fa;
                    border: 2px solid #e9ecef;
                    border-radius: 6px;
                    padding: 15px;
                    min-height: 60px;
                    line-height: 1.5;
                }}
                .zone-list {{
                    background: #f8f9fa;
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    padding: 15px;
                }}
                .zone-list ul {{
                    margin: 0;
                    padding-left: 20px;
                }}
                .zone-list li {{
                    margin-bottom: 5px;
                    color: #495057;
                }}
                .totals-info {{
                    text-align: right;
                    color: #6c757d;
                    font-size: 10px;
                    margin-top: 10px;
                    font-style: italic;
                }}
                @media print {{
                    body {{ margin: 10px; }}
                    .section {{ page-break-inside: avoid; }}
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="title">{title}</div>
                <div class="subtitle">সিটি রিপোর্ট সারসংক্ষেপ</div>
                <div class="report-info">রিপোর্টের ধরন: {city_data['report_type']} | বছর: {city_data['year']}</div>
            </div>

            <!-- Header Section -->
            <div class="section">
                <div class="section-title">📋 মূল তথ্য</div>
                <div class="header-grid">
        """

        # Add header data in grid format matching the web template
        header_items = [
            ("responsible_name", "দায়িত্বশীলের নাম"),
            ("thana", "থানা"),
            ("ward", "ওয়ার্ড"),
            ("total_muallima", "মোট মুয়াল্লিমা সংখ্যা"),
            ("muallima_increase", "মুয়াল্লিমা বৃদ্ধি"),
            ("muallima_decrease", "মুয়াল্লিমা হ্রাস"),
            ("certified_muallima", "প্রশিক্ষিত মুয়াল্লিমা"),
            ("certified_muallima_taking_classes", "প্রশিক্ষিত মুয়াল্লিমা (ক্লাস নিচ্ছেন)"),
            ("trained_muallima", "ট্রেইনিং প্রাপ্ত মুয়াল্লিমা"),
            ("trained_muallima_taking_classes", "ট্রেইনিং প্রাপ্ত মুয়াল্লিমা (ক্লাস নিচ্ছেন)"),
            ("total_unit", "মোট ইউনিট"),
            ("units_with_muallima", "মুয়াল্লিমা সহ ইউনিট"),
        ]

        for field, label in header_items:
            value = city_data["city_summary"].get(field)
            display_value = value if value is not None else "N/A"
            html_content += f"""
                    <div class="header-item">
                        <span class="header-label">{label}:</span>
                        <span class="header-value">{display_value}</span>
                    </div>
                """

        html_content += """
                </div>
            </div>

            <!-- Courses Section -->
            <div class="section">
                <div class="section-title">📚 গ্রুপ / কোর্স রিপোর্ট</div>
                <table>
                    <thead>
                        <tr>
                            <th rowspan="2">বিভাগ/ধরন</th>
                            <th colspan="3">গ্রুপ / কোর্স</th>
                            <th rowspan="2">অধিবেশন সংখ্যা</th>
                            <th rowspan="2">শিক্ষার্থী সংখ্যা</th>
                            <th rowspan="2">উপস্থিতি সংখ্যা</th>
                            <th colspan="4">শিক্ষার্থী অবস্থান</th>
                            <th rowspan="2">কতজন নিয়ে সমাপ্ত</th>
                            <th rowspan="2">সহীহ শিখেছেন কতজন</th>
                        </tr>
                        <tr class="sub-header">
                            <th>সংখ্যা</th>
                            <th>বৃদ্ধি</th>
                            <th>ঘাটতি</th>
                            <th>বোর্ডে</th>
                            <th>কায়দায়</th>
                            <th>আমপারা</th>
                            <th>কুরআন</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        for course in city_data["city_courses"]:
            html_content += f"""
                        <tr>
                            <td class="category">{course['category']}</td>
                            <td class="number">{course['number']}</td>
                            <td class="number">{course['increase']}</td>
                            <td class="number">{course['decrease']}</td>
                            <td class="number">{course['sessions']}</td>
                            <td class="number">{course['students']}</td>
                            <td class="number">{course['attendance']}</td>
                            <td class="number">{course['status_board']}</td>
                            <td class="number">{course['status_qayda']}</td>
                            <td class="number">{course['status_ampara']}</td>
                            <td class="number">{course['status_quran']}</td>
                            <td class="number">{course['completed']}</td>
                            <td class="number">{course['correctly_learned']}</td>
                        </tr>
            """

        html_content += """
                    </tbody>
                </table>
            </div>

            <!-- Organizational Section -->
            <div class="section">
                <div class="section-title">🏢 দাওয়াত ও সংগঠন</div>
                <table>
                    <thead>
                        <tr>
                            <th>দাওয়াত ও সংগঠন</th>
                            <th>সংখ্যা</th>
                            <th>বৃদ্ধি</th>
                            <th>পরিমাণ</th>
                            <th>মন্তব্য</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        for org in city_data["city_organizational"]:
            html_content += f"""
                        <tr>
                            <td class="category">{org['category']}</td>
                            <td class="number">{org['number']}</td>
                            <td class="number">{org['increase']}</td>
                            <td class="number">{org['amount']}</td>
                            <td>{org['comments'] if org['comments'] is not None else ''}</td>
                        </tr>
            """

        html_content += """
                    </tbody>
                </table>
            </div>

            <!-- Personal Section -->
            <div class="section">
                <div class="section-title">ব্যক্তিগত উদ্যোগে তা'লীমুল কুরআন</div>
                <table>
                    <thead>
                        <tr>
                            <th rowspan="2">ব্যক্তিগত উদ্যোগে তা'লীমুল কুরআন</th>
                            <th rowspan="2">কতজন শিখাচ্ছেন</th>
                            <th rowspan="2" style="border-right: 4px solid #3498db;">কতজনকে শিখাচ্ছেন</th>
                            <th rowspan="2">কতজন ওয়ালামাকে দাওয়াত দিয়েছেন</th>
                            <th colspan="4">দাওয়াত প্রাপ্ত ওয়ালামার মধ্যে</th>
                        </tr>
                        <tr class="sub-header">
                            <th>সহযোগী হয়েছেন</th>
                            <th>সক্রিয় সহযোগী হয়েছেন</th>
                            <th>কর্মী হয়েছেন</th>
                            <th>রুকন হয়েছেন</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        # Calculate totals for personal section
        teaching_total = sum(p["teaching"] for p in city_data["city_personal"])
        learning_total = sum(p["learning"] for p in city_data["city_personal"])
        olama_invited_total = sum(
            p["olama_invited"] for p in city_data["city_personal"]
        )
        became_shohojogi_total = sum(
            p["became_shohojogi"] for p in city_data["city_personal"]
        )
        became_sokrio_shohojogi_total = sum(
            p["became_sokrio_shohojogi"] for p in city_data["city_personal"]
        )
        became_kormi_total = sum(p["became_kormi"] for p in city_data["city_personal"])
        became_rukon_total = sum(p["became_rukon"] for p in city_data["city_personal"])

        for personal in city_data["city_personal"]:
            html_content += f"""
                        <tr>
                            <td class="category">{personal['category']}</td>
                            <td class="number">{personal['teaching']}</td>
                            <td class="number" style="border-right: 4px solid #3498db;">{personal['learning']}</td>
                            <td class="number">{personal['olama_invited']}</td>
                            <td class="number">{personal['became_shohojogi']}</td>
                            <td class="number">{personal['became_sokrio_shohojogi']}</td>
                            <td class="number">{personal['became_kormi']}</td>
                            <td class="number">{personal['became_rukon']}</td>
                        </tr>
            """

        # Add totals row for personal section
        html_content += f"""
                        <tr class="totals">
                            <td class="category">মোট</td>
                            <td class="number">{teaching_total}</td>
                            <td class="number" style="border-right: 4px solid #1a5490;">{learning_total}</td>
                            <td class="number">{olama_invited_total}</td>
                            <td class="number">{became_shohojogi_total}</td>
                            <td class="number">{became_sokrio_shohojogi_total}</td>
                            <td class="number">{became_kormi_total}</td>
                            <td class="number">{became_rukon_total}</td>
                        </tr>
        """

        html_content += """
                    </tbody>
                </table>
            </div>

            <!-- Meetings Section -->
            <div class="section">
                <div class="section-title">🤝 বৈঠকসমূহ</div>
                <table>
                    <thead>
                        <tr>
                            <th>বৈঠকসমূহ</th>
                            <th>মহানগরীর কতটি</th>
                            <th>মহানগরী গড় উপস্থিতি</th>
                            <th>থানা কতটি</th>
                            <th>থানা গড় উপস্থিতি</th>
                            <th>ওয়ার্ড কতটি</th>
                            <th>ওয়ার্ড গড় উপস্থিতি</th>
                            <th>মন্তব্য</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        for meeting in city_data["city_meetings"]:
            html_content += f"""
                        <tr>
                            <td class="category">{meeting['category']}</td>
                            <td class="number">{meeting['city_count']}</td>
                            <td class="number">{meeting['city_avg_attendance']}</td>
                            <td class="number">{meeting['thana_count']}</td>
                            <td class="number">{meeting['thana_avg_attendance']}</td>
                            <td class="number">{meeting['ward_count']}</td>
                            <td class="number">{meeting['ward_avg_attendance']}</td>
                            <td>{meeting['comments'] if meeting['comments'] is not None else ''}</td>
                        </tr>
            """

        html_content += """
                    </tbody>
                </table>
            </div>

            <!-- Extras Section -->
            <div class="section">
                <div class="section-title">📊 মক্তব ও সফর রিপোর্ট</div>
                <table>
                    <thead>
                        <tr>
                            <th>বিষয়</th>
                            <th>সংখ্যা</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        for extra in city_data["city_extras"]:
            html_content += f"""
                        <tr>
                            <td class="category">{extra['category']}</td>
                            <td class="number">{extra['number']}</td>
                        </tr>
            """

        html_content += f"""
                    </tbody>
                </table>
            </div>

            <!-- Comments Section -->
            <div class="section">
                <div class="section-title">💬 মন্তব্য</div>
                <div class="comments-box">
                    {city_data['city_comments']['comment'] if city_data['city_comments']['comment'] else 'কোনো মন্তব্য নেই'}
                </div>
            </div>
        """

        html_content += """
        </body>
        </html>
        """

        # Generate PDF using Playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()

            # Set content and wait for fonts to load
            page.set_content(html_content)
            page.wait_for_load_state("networkidle")

            # Generate PDF with high quality settings
            pdf_bytes = page.pdf(
                format="A4",
                print_background=True,
                margin={
                    "top": "0.5in",
                    "bottom": "0.5in",
                    "left": "0.5in",
                    "right": "0.5in",
                },
            )

            browser.close()

            # Return PDF as response
            output = io.BytesIO(pdf_bytes)
            output.seek(0)

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype="application/pdf",
            )

    except Exception as e:
        print(f"[DEBUG] City Report PDF generation failed: {e}")
        flash(f"PDF generation failed: {str(e)}", "error")
        return redirect(url_for("city_report_page"))


def generate_pdf_with_playwright(reports, title, filename):
    """Generate PDF using Playwright for perfect Bengali support"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise Exception(
            "Playwright not installed. Install with: pip install playwright"
        )

    # Create comprehensive HTML content with all report sections
    html_content = f"""
    <!DOCTYPE html>
    <html lang="bn">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{title}</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+Bengali:wght@400;700&display=swap');

            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}

            body {{
                font-family: 'Noto Sans Bengali', Arial, sans-serif;
                font-size: 12px;
                line-height: 1.4;
                color: #333;
                margin: 15px;
            }}

            .header {{
                text-align: center;
                margin-bottom: 25px;
                border-bottom: 2px solid #333;
                padding-bottom: 15px;
            }}

            .title {{
                font-size: 18px;
                font-weight: bold;
                margin-bottom: 5px;
                color: #2c3e50;
            }}

            .subtitle {{
                font-size: 14px;
                color: #7f8c8d;
            }}

            .section {{
                margin-bottom: 20px;
                page-break-inside: avoid;
            }}

            .section-title {{
                font-size: 16px;
                font-weight: bold;
                background: linear-gradient(135deg, #3498db, #2980b9);
                color: white;
                padding: 8px 12px;
                margin-bottom: 10px;
                border-radius: 4px;
                text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
            }}

            table {{
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 15px;
                font-size: 11px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            }}

            th {{
                background: linear-gradient(135deg, #34495e, #2c3e50);
                color: white;
                padding: 10px 8px;
                text-align: center;
                font-weight: bold;
                border: 1px solid #2c3e50;
                text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
            }}

            td {{
                padding: 8px;
                border: 1px solid #bdc3c7;
                text-align: center;
                background-color: #ffffff;
            }}

            tr:nth-child(even) td {{
                background-color: #f8f9fa;
            }}

            tr:hover td {{
                background-color: #e8f4fd;
            }}

            .field-name {{
                font-weight: bold;
                background-color: #ecf0f1 !important;
                text-align: left;
                color: #2c3e50;
            }}

            .number-cell {{
                font-weight: bold;
                color: #27ae60;
            }}

            .timestamp {{
                margin-top: 20px;
                text-align: center;
                font-size: 10px;
                color: #7f8c8d;
                border-top: 1px solid #bdc3c7;
                padding-top: 10px;
            }}

            .page-break {{
                page-break-before: always;
            }}

            @media print {{
                body {{ margin: 10px; }}
                .section {{ page-break-inside: avoid; }}
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="title">{title}</div>
            <div class="subtitle">রিপোর্ট সারসংক্ষেপ</div>
        </div>
    """

    # Process each report
    for report_index, report in enumerate(reports):
        if report_index > 0:
            html_content += '<div class="page-break"></div>'

        # Header Section
        if report.header:
            html_content += """
            <div class="section">
                <div class="section-title">📋 হেডার তথ্য (Header Information)</div>
                <table>
                    <thead>
                        <tr>
                            <th style="width: 40%;">ক্ষেত্র</th>
                            <th style="width: 60%;">মান</th>
                        </tr>
                    </thead>
                    <tbody>"""

            header_data = [
                ("দায়িত্বশীলের নাম", report.header.responsible_name or ""),
                ("থানা", str(report.header.thana or "")),
                ("ওয়ার্ড", str(report.header.ward or "")),
                ("মোট মুয়াল্লিমা", str(report.header.total_muallima or 0)),
                ("মুয়াল্লিমা বৃদ্ধি", str(report.header.muallima_increase or 0)),
                ("মুয়াল্লিমা হ্রাস", str(report.header.muallima_decrease or 0)),
                ("সার্টিফিকেটধারী মুয়াল্লিমা", str(report.header.certified_muallima or 0)),
                (
                    "সার্টিফিকেটধারী মুয়াল্লিমা ক্লাস নিচ্ছেন",
                    str(report.header.certified_muallima_taking_classes or 0),
                ),
                ("প্রশিক্ষিত মুয়াল্লিমা", str(report.header.trained_muallima or 0)),
                (
                    "প্রশিক্ষিত মুয়াল্লিমা ক্লাস নিচ্ছেন",
                    str(report.header.trained_muallima_taking_classes or 0),
                ),
                ("মোট ইউনিট", str(report.header.total_unit or 0)),
                ("মুয়াল্লিমা সহ ইউনিট", str(report.header.units_with_muallima or 0)),
            ]

            for field, value in header_data:
                html_content += f"""
                        <tr>
                            <td class="field-name">{field}</td>
                            <td class="number-cell">{value}</td>
                        </tr>"""

            html_content += """
                    </tbody>
                </table>
            </div>"""

        # Courses Section
        if report.courses:
            html_content += """
            <div class="section">
                <div class="section-title">📚 গ্রুপ / কোর্স রিপোর্ট</div>
                <table>
                    <thead>
                        <tr>
                            <th rowspan="2" style="width: 15%;">বিভাগ/ধরন</th>
                            <th colspan="3" style="border-bottom: 1px solid #bdc3c7;">গ্রুপ / কোর্স</th>
                            <th rowspan="2">অধিবেশন সংখ্যা</th>
                            <th rowspan="2">শিক্ষার্থী সংখ্যা</th>
                            <th rowspan="2">উপস্থিতি সংখ্যা</th>
                            <th colspan="4" style="border-bottom: 1px solid #bdc3c7;">শিক্ষার্থী অবস্থান</th>
                            <th rowspan="2">কতজন নিয়ে সমাপ্ত</th>
                            <th rowspan="2">সহীহ শিখেছেন কতজন</th>
                        </tr>
                        <tr>
                            <th>সংখ্যা</th>
                            <th>বৃদ্ধি</th>
                            <th>ঘাটতি</th>
                            <th>বোর্ডে</th>
                            <th>কায়দায়</th>
                            <th>আমপারা</th>
                            <th>কুরআন</th>
                        </tr>
                    </thead>
                    <tbody>"""

            course_categories = [
                "বিশিষ্টদের",
                "সাধারণদের",
                "কর্মীদের",
                "ইউনিট সভানেত্রী",
                "অগ্রসরদের",
                "শিশু- তা'লিমুল কুরআন",
                "নিরক্ষর- তা'লিমুস সলাত",
            ]

            for category in course_categories:
                # Find the course row for this category
                course_row = next(
                    (c for c in report.courses if c.category == category), None
                )

                html_content += f"""
                        <tr>
                            <td class="field-name">{category}</td>
                            <td class="number-cell">{course_row.number if course_row and course_row.number is not None else 0}</td>
                            <td class="number-cell">{course_row.increase if course_row and course_row.increase is not None else 0}</td>
                            <td class="number-cell">{course_row.decrease if course_row and course_row.decrease is not None else 0}</td>
                            <td class="number-cell">{course_row.sessions if course_row and course_row.sessions is not None else 0}</td>
                            <td class="number-cell">{course_row.students if course_row and course_row.students is not None else 0}</td>
                            <td class="number-cell">{course_row.attendance if course_row and course_row.attendance is not None else 0}</td>
                            <td class="number-cell">{course_row.status_board if course_row and course_row.status_board is not None else 0}</td>
                            <td class="number-cell">{course_row.status_qayda if course_row and course_row.status_qayda is not None else 0}</td>
                            <td class="number-cell">{course_row.status_ampara if course_row and course_row.status_ampara is not None else 0}</td>
                            <td class="number-cell">{course_row.status_quran if course_row and course_row.status_quran is not None else 0}</td>
                            <td class="number-cell">{course_row.completed if course_row and course_row.completed is not None else 0}</td>
                            <td class="number-cell">{course_row.correctly_learned if course_row and course_row.correctly_learned is not None else 0}</td>
                        </tr>"""

            html_content += """
                    </tbody>
                </table>
            </div>"""

        # Organizational Section
        if report.organizational:
            html_content += """
            <div class="section">
                <div class="section-title">🏢 দাওয়াত ও সংগঠন</div>
                <table>
                    <thead>
                        <tr>
                            <th style="width: 40%;">দাওয়াত ও সংগঠন</th>
                            <th>সংখ্যা</th>
                            <th>বৃদ্ধি</th>
                            <th>পরিমাণ</th>
                            <th>মন্তব্য</th>
                        </tr>
                    </thead>
                    <tbody>"""

            org_categories = [
                "দাওয়াত দান",
                "কতজন ইসলামের আদর্শ মেনে চলার চেষ্টা করছেন",
                "সহযোগী হয়েছে",
                "সম্মতি দিয়েছেন",
                "সক্রিয় সহযোগী",
                "কর্মী",
                "রুকন",
                "দাওয়াতী ইউনিট",
                "ইউনিট",
                "সূধী",
                "এককালীন",
                "জনশক্তির সহীহ্ কুরআন তিলাওয়াত অনুশীলনী (মাশক)",
                "বই বিলি",
                "বই বিক্রি",
            ]

            for category in org_categories:
                # Find the organizational row for this category
                org_row = next(
                    (o for o in report.organizational if o.category == category), None
                )

                html_content += f"""
                        <tr>
                            <td class="field-name">{category}</td>
                            <td class="number-cell">{org_row.number if org_row and org_row.number is not None else 0}</td>
                            <td class="number-cell">{org_row.increase if org_row and org_row.increase is not None else 0}</td>
                            <td class="number-cell">{org_row.amount if org_row and org_row.amount is not None else 0}</td>
                            <td>{org_row.comments if org_row and org_row.comments is not None else ""}</td>
                        </tr>"""

            html_content += """
                    </tbody>
                </table>
            </div>"""

        # Personal Section
        if report.personal:
            html_content += """
            <div class="section">
                <div class="section-title">👤 ব্যক্তিগত উদ্যোগে তা'লীমুল কুরআন</div>
                <table>
                    <thead>
                        <tr>
                            <th rowspan="2" style="width: 20%;">ব্যক্তিগত উদ্যোগে তা'লীমুল কুরআন</th>
                            <th rowspan="2">কতজন শিখাচ্ছেন</th>
                            <th rowspan="2">কতজনকে শিখাচ্ছেন</th>
                            <th rowspan="2">কতজন ওয়ালামাকে দাওয়াত দিয়েছেন</th>
                            <th colspan="4" style="border-bottom: 1px solid #bdc3c7;">দাওয়াত প্রাপ্ত ওয়ালামার মধ্যে</th>
                        </tr>
                        <tr>
                            <th>সহযোগী হয়েছেন</th>
                            <th>সক্রিয় সহযোগী হয়েছেন</th>
                            <th>কর্মী হয়েছেন</th>
                            <th>রুকন হয়েছেন</th>
                        </tr>
                    </thead>
                    <tbody>"""

            personal_categories = ["রুকন", "কর্মী", "সক্রিয় সহযোগী"]

            # Calculate totals
            teaching_total = 0
            learning_total = 0
            olama_invited_total = 0
            became_shohojogi_total = 0
            became_sokrio_shohojogi_total = 0
            became_kormi_total = 0
            became_rukon_total = 0

            for category in personal_categories:
                # Find the personal row for this category
                personal_row = next(
                    (p for p in report.personal if p.category == category), None
                )

                teaching_val = (
                    personal_row.teaching
                    if personal_row and personal_row.teaching is not None
                    else 0
                )
                learning_val = (
                    personal_row.learning
                    if personal_row and personal_row.learning is not None
                    else 0
                )
                olama_invited_val = (
                    personal_row.olama_invited
                    if personal_row and personal_row.olama_invited is not None
                    else 0
                )
                became_shohojogi_val = (
                    personal_row.became_shohojogi
                    if personal_row and personal_row.became_shohojogi is not None
                    else 0
                )
                became_sokrio_shohojogi_val = (
                    personal_row.became_sokrio_shohojogi
                    if personal_row and personal_row.became_sokrio_shohojogi is not None
                    else 0
                )
                became_kormi_val = (
                    personal_row.became_kormi
                    if personal_row and personal_row.became_kormi is not None
                    else 0
                )
                became_rukon_val = (
                    personal_row.became_rukon
                    if personal_row and personal_row.became_rukon is not None
                    else 0
                )

                # Add to totals
                teaching_total += teaching_val
                learning_total += learning_val
                olama_invited_total += olama_invited_val
                became_shohojogi_total += became_shohojogi_val
                became_sokrio_shohojogi_total += became_sokrio_shohojogi_val
                became_kormi_total += became_kormi_val
                became_rukon_total += became_rukon_val

                html_content += f"""
                        <tr>
                            <td class="field-name">{category}</td>
                            <td class="number-cell">{teaching_val}</td>
                            <td class="number-cell">{learning_val}</td>
                            <td class="number-cell">{olama_invited_val}</td>
                            <td class="number-cell">{became_shohojogi_val}</td>
                            <td class="number-cell">{became_sokrio_shohojogi_val}</td>
                            <td class="number-cell">{became_kormi_val}</td>
                            <td class="number-cell">{became_rukon_val}</td>
                        </tr>"""

            # Add totals row
            html_content += f"""
                        <tr style="background-color: #e0f2fe; font-weight: bold;">
                            <td class="field-name">মোট</td>
                            <td class="number-cell">{teaching_total}</td>
                            <td class="number-cell">{learning_total}</td>
                            <td class="number-cell">{olama_invited_total}</td>
                            <td class="number-cell">{became_shohojogi_total}</td>
                            <td class="number-cell">{became_sokrio_shohojogi_total}</td>
                            <td class="number-cell">{became_kormi_total}</td>
                            <td class="number-cell">{became_rukon_total}</td>
                        </tr>"""

            html_content += """
                    </tbody>
                </table>
            </div>"""

        # Meetings Section
        if report.meetings:
            html_content += """
            <div class="section">
                <div class="section-title">🤝 বৈঠকসমূহ</div>
                <table>
                    <thead>
                        <tr>
                            <th style="width: 25%;">বৈঠকসমূহ</th>
                            <th>মহানগরীর কতটি</th>
                            <th>মহানগরী গড় উপস্থিতি</th>
                            <th>থানা কতটি</th>
                            <th>থানা গড় উপস্থিতি</th>
                            <th>ওয়ার্ড কতটি</th>
                            <th>ওয়ার্ড গড় উপস্থিতি</th>
                            <th>মন্তব্য</th>
                        </tr>
                    </thead>
                    <tbody>"""

            meeting_categories = [
                "কমিটি বৈঠক হয়েছে",
                "মুয়াল্লিমাদের নিয়ে বৈঠক",
                "Committee Orientation",
                "Muallima Orientation",
            ]

            for category in meeting_categories:
                # Find the meeting row for this category
                meeting_row = next(
                    (m for m in report.meetings if m.category == category), None
                )

                html_content += f"""
                        <tr>
                            <td class="field-name">{category}</td>
                            <td class="number-cell">{meeting_row.city_count if meeting_row and meeting_row.city_count is not None else 0}</td>
                            <td class="number-cell">{meeting_row.city_avg_attendance if meeting_row and meeting_row.city_avg_attendance is not None else 0}</td>
                            <td class="number-cell">{meeting_row.thana_count if meeting_row and meeting_row.thana_count is not None else 0}</td>
                            <td class="number-cell">{meeting_row.thana_avg_attendance if meeting_row and meeting_row.thana_avg_attendance is not None else 0}</td>
                            <td class="number-cell">{meeting_row.ward_count if meeting_row and meeting_row.ward_count is not None else 0}</td>
                            <td class="number-cell">{meeting_row.ward_avg_attendance if meeting_row and meeting_row.ward_avg_attendance is not None else 0}</td>
                            <td>{meeting_row.comments if meeting_row and meeting_row.comments is not None else ""}</td>
                        </tr>"""

            html_content += """
                    </tbody>
                </table>
            </div>"""

        # Extras Section
        if report.extras:
            html_content += """
            <div class="section">
                <div class="section-title">➕ মক্তব ও সফর রিপোর্ট</div>
                <table>
                    <thead>
                        <tr>
                            <th style="width: 70%;">বিষয়</th>
                            <th style="width: 30%;">সংখ্যা</th>
                        </tr>
                    </thead>
                    <tbody>"""

            extra_categories = [
                "মক্তব সংখ্যা",
                "মক্তব বৃদ্ধি",
                "মহানগরী পরিচালিত",
                "স্থানীয়ভাবে পরিচালিত",
                "মহানগরীর সফর",
                "থানা কমিটির সফর",
                "থানা প্রতিনিধির সফর",
                "ওয়ার্ড প্রতিনিধির সফর",
            ]

            for category in extra_categories:
                # Find the extras row for this category
                extra_row = next(
                    (e for e in report.extras if e.category == category), None
                )

                html_content += f"""
                        <tr>
                            <td class="field-name">{category}</td>
                            <td class="number-cell">{extra_row.number if extra_row and extra_row.number is not None else 0}</td>
                        </tr>"""

            html_content += """
                    </tbody>
                </table>
            </div>"""

        # Comments Section
        if report.comments and report.comments.comment:
            html_content += f"""
            <div class="section">
                <div class="section-title">💬 মন্তব্য</div>
                <table>
                    <tbody>
                        <tr>
                            <td style="padding: 15px; text-align: left; background-color: #f8f9fa;">
                                {report.comments.comment}
                            </td>
                        </tr>
                    </tbody>
                </table>
            </div>"""

    # Add timestamp
    html_content += f"""
        <div class="timestamp">
            রিপোর্ট তৈরি হয়েছে: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 
            সিস্টেম: রিপোর্ট সাবমিশন সিস্টেম
        </div>
    </body>
    </html>
    """

    # Generate PDF using Playwright
    print("[DEBUG] Starting Playwright PDF generation...")
    with sync_playwright() as p:
        try:
            print("[DEBUG] Launching browser...")
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            print("[DEBUG] Browser launched, setting content...")

            # Set content and wait for fonts to load
            page.set_content(html_content, wait_until="networkidle")
            print("[DEBUG] Content set, waiting for fonts...")

            # Wait a bit for Google Fonts to load
            page.wait_for_timeout(2000)
            print("[DEBUG] Generating PDF...")

            # Generate PDF with high quality settings
            pdf_bytes = page.pdf(
                format="A4",
                margin={
                    "top": "0.5in",
                    "right": "0.5in",
                    "bottom": "0.5in",
                    "left": "0.5in",
                },
                print_background=True,
                prefer_css_page_size=True,
            )

            print(f"[DEBUG] PDF generated successfully, size: {len(pdf_bytes)} bytes")
            browser.close()

            # Return as downloadable file
            import io

            output = io.BytesIO(pdf_bytes)
            print("[DEBUG] Returning PDF file...")
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype="application/pdf",
            )

        except Exception as e:
            print(f"[DEBUG] Playwright PDF generation failed: {e}")
            browser.close()
        raise


def get_reports_for_period(zone_id, report_type, month, year):

    # Convert month name to number if needed
    month_name_to_number = {
        "জানুয়ারি": 1,
        "ফেব্রুয়ারি": 2,
        "মার্চ": 3,
        "এপ্রিল": 4,
        "মে": 5,
        "জুন": 6,
        "জুলাই": 7,
        "আগস্ট": 8,
        "সেপ্টেম্বর": 9,
        "অক্টোবর": 10,
        "নভেম্বর": 11,
        "ডিসেম্বর": 12,
    }

    if report_type == "মাসিক":
        # Convert month name to number if it's a string
        if isinstance(month, str) and month in month_name_to_number:
            month_nums = [month_name_to_number[month]]
        else:
            month_nums = [int(month)]
    elif report_type == "ত্রৈমাসিক":
        month_nums = [1, 2, 3]
    elif report_type == "ষান্মাসিক":
        month_nums = [1, 2, 3, 4, 5, 6]
    elif report_type == "নয়-মাসিক":
        month_nums = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    elif report_type == "বার্ষিক":
        month_nums = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    else:
        # Default to monthly
        if isinstance(month, str) and month in month_name_to_number:
            month_nums = [month_name_to_number[month]]
        else:
            month_nums = [int(month)]

    reports = []
    for m in month_nums:
        report = Report.query.filter_by(zone_id=zone_id, month=m, year=year).first()
        if report:
            reports.append(report)

    return reports


# --- Help Page ---


@app.route("/help")
def help_page():
    return render_template("help.html")


# --- Error Handlers ---


@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


# --- Main ---

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 5000)),
        debug=os.environ.get("FLASK_DEBUG", "false").lower() == "true",
    )
