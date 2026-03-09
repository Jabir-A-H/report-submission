import os

FILE_PATH = r"f:\\WebDev\\report-submission\\services\\excel_export.py"

NEW_EXCEL_EXPORT = '''import io
import pandas as pd
from utils import COURSE_CATEGORIES, ORG_CATEGORIES, PERSONAL_CATEGORIES, MEETING_CATEGORIES, EXTRA_CATEGORIES

def _safe_get(dictionary, key, default=0):
    val = dictionary.get(key)
    return val if val is not None else default

def generate_excel_file(reports, report_type):
    from schemas import ReportSchema
    
    # Serialize the reports into JSON via Marshmallow
    serialized_reports = ReportSchema(many=True).dump(reports)
    
    data = []
    
    for report_index, report in enumerate(serialized_reports):
        # Add zone separator
        if report_index > 0:
            data.extend([[""] * 7, [""] * 7])  # Empty rows

        # Zone Header
        zone_name = report.get("zone", {}).get("name", "") if report.get("zone") else ""
        if zone_name:
            data.extend([[f"জোন: {zone_name}"] + [""] * 6, [""] * 7])

        # Header section
        header = report.get("header") or {}
        if header:
            data.append(["বিভাগ", "ক্ষেত্র", "মান", "", "", "", ""])
            data.append(["হেডার তথ্য", "দায়িত্বশীলের নাম", header.get("responsible_name") or "", "", "", "", ""])
            data.append(["হেডার তথ্য", "থানা", str(header.get("thana") or ""), "", "", "", ""])
            data.append(["হেডার তথ্য", "ওয়ার্ড", str(header.get("ward") or ""), "", "", "", ""])
            data.append(["হেডার তথ্য", "মোট মুয়াল্লিমা", str(_safe_get(header, "total_muallima")), "", "", "", ""])
            data.append(["হেডার তথ্য", "মুয়াল্লিমা বৃদ্ধি", str(_safe_get(header, "muallima_increase")), "", "", "", ""])
            data.append(["হেডার তথ্য", "মুয়াল্লিমা হ্রাস", str(_safe_get(header, "muallima_decrease")), "", "", "", ""])
            data.append(["হেডার তথ্য", "সার্টিফিকেটধারী মুয়াল্লিমা", str(_safe_get(header, "certified_muallima")), "", "", "", ""])
            data.append(["হেডার তথ্য", "সার্টিফিকেটধারী মুয়াল্লিমা ক্লাস নিচ্ছেন", str(_safe_get(header, "certified_muallima_taking_classes")), "", "", "", ""])
            data.append(["হেডার তথ্য", "প্রশিক্ষিত মুয়াল্লিমা", str(_safe_get(header, "trained_muallima")), "", "", "", ""])
            data.append(["হেডার তথ্য", "প্রশিক্ষিত মুয়াল্লিমা ক্লাস নিচ্ছেন", str(_safe_get(header, "trained_muallima_taking_classes")), "", "", "", ""])
            data.append(["হেডার তথ্য", "মোট ইউনিট", str(_safe_get(header, "total_unit")), "", "", "", ""])
            data.append(["হেডার তথ্য", "মুয়াল্লিমা সহ ইউনিট", str(_safe_get(header, "units_with_muallima")), "", "", "", ""])
            data.append([""] * 7)

        # Courses section
        courses = report.get("courses") or []
        if courses:
            data.append(["বিভাগ", "গ্রুপ/কোর্স", "সংখ্যা", "বৃদ্ধি", "ঘাটতি", "অধিবেশন", "শিক্ষার্থী"])
            for category in COURSE_CATEGORIES:
                c = next((item for item in courses if item.get("category") == category), {})
                data.append([
                    "গ্রুপ/কোর্স", category,
                    _safe_get(c, "number"), _safe_get(c, "increase"), _safe_get(c, "decrease"),
                    _safe_get(c, "sessions"), _safe_get(c, "students")
                ])
            data.append([""] * 7)

        # Organizational section
        organizational = report.get("organizational") or []
        if organizational:
            data.append(["বিভাগ", "দাওয়াত ও সংগঠন", "সংখ্যা", "বৃদ্ধি", "পরিমাণ", "মন্তব্য", ""])
            for category in ORG_CATEGORIES:
                o = next((item for item in organizational if item.get("category") == category), {})
                data.append([
                    "দাওয়াত ও সংগঠন", category,
                    _safe_get(o, "number"), _safe_get(o, "increase"), _safe_get(o, "amount"),
                    o.get("comments") or "", ""
                ])
            data.append([""] * 7)

        # Personal section
        personal = report.get("personal") or []
        if personal:
            data.append(["বিভাগ", "ব্যক্তিগত উদ্যোগে তা'লীমুল কুরআন", "কতজন শিখাচ্ছেন", "কতজনকে শিখাচ্ছেন", "ওয়ালামাকে দাওয়াত", "সহযোগী হয়েছেন", "কর্মী হয়েছেন"])
            for category in PERSONAL_CATEGORIES:
                p = next((item for item in personal if item.get("category") == category), {})
                data.append([
                    "ব্যক্তিগত", category,
                    _safe_get(p, "teaching"), _safe_get(p, "learning"), _safe_get(p, "olama_invited"),
                    _safe_get(p, "became_shohojogi"), _safe_get(p, "became_kormi")
                ])
            data.append([""] * 7)

        # Meetings section
        meetings = report.get("meetings") or []
        if meetings:
            data.append(["বিভাগ", "বৈঠকসমূহ", "মহানগরী কতটি", "মহানগরী গড় উপস্থিতি", "থানা কতটি", "থানা গড় উপস্থিতি", "মন্তব্য"])
            for category in MEETING_CATEGORIES:
                m = next((item for item in meetings if item.get("category") == category), {})
                data.append([
                    "বৈঠকসমূহ", category,
                    _safe_get(m, "city_count"), _safe_get(m, "city_avg_attendance"),
                    _safe_get(m, "thana_count"), _safe_get(m, "thana_avg_attendance"),
                    m.get("comments") or ""
                ])
            data.append([""] * 7)

        # Extras section
        extras = report.get("extras") or []
        if extras:
            data.append(["বিভাগ", "মক্তব ও সফর রিপোর্ট", "সংখ্যা", "", "", "", ""])
            for category in EXTRA_CATEGORIES:
                e = next((item for item in extras if item.get("category") == category), {})
                data.append([
                    "মক্তব ও সফর", category,
                    _safe_get(e, "number"), "", "", "", ""
                ])
            data.append([""] * 7)

        # Comments section
        comments = report.get("comments") or {}
        if comments and comments.get("comment"):
            data.append(["বিভাগ", "মন্তব্য", "", "", "", "", ""])
            data.append(["মন্তব্য", comments.get("comment"), "", "", "", "", ""])
            data.append([""] * 7)

    # Create DataFrame with flexible column structure
    headers = ["বিভাগ", "বিষয়", "কলাম ১", "কলাম ২", "কলাম ৩", "কলাম ৪", "কলাম ৫"]
    df = pd.DataFrame(data, columns=headers)

    # Create Excel file in memory
    output = io.BytesIO()

    # Download Tiro Bangla font data for Excel
    bengali_font_name = "Times New Roman"  # Default fallback
    try:
        import urllib.request
        font_url = "https://fonts.gstatic.com/s/tirobangla/v6/IuaHaJaHVsk2rGGByzUFTJrmwLs.ttf"
        with urllib.request.urlopen(font_url, timeout=2) as response:
            bengali_font_name = "Tiro Bangla"
    except:
        pass

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Report", index=False)

        from openpyxl.styles import Font, PatternFill, Border, Side

        workbook = writer.book
        worksheet = writer.sheets["Report"]

        bengali_font = Font(name=bengali_font_name, size=11)
        header_font = Font(name=bengali_font_name, size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = bengali_font
                cell.border = thin_border

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return output
'''

with open(FILE_PATH, 'w', encoding='utf-8') as f:
    f.write(NEW_EXCEL_EXPORT)

print("Report excel generator decoupled with Marshmallow!")
